"""
╔══════════════════════════════════════════════════════════════════════╗
║           GOOFY SCREENER — PHASE 3                                   ║
║           Multi-Market Autonomous Strategy Screener                  ║
║                                                                      ║
║   Markets:   🇺🇸 US  |  🇦🇺 ASX  |  🇯🇵 Japan (JPX)               ║
║   Strategies: MA Crossover, RSI, Bollinger Bands, MACD, Mean Rev    ║
║   Output:    Tiered ranking (S/A/B/Skip) + Excel multi-tab report   ║
║                                                                      ║
║   Run:  python goofy_screener_phase3.py                              ║
║   Args: --market US        (default: ALL)                            ║
║         --market ASX                                                 ║
║         --market JPX                                                 ║
║         --market ALL                                                 ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import yfinance as yf
import numpy as np
import pandas as pd
import datetime as dt
import os, sys, warnings, argparse
warnings.filterwarnings("ignore")

# ── Optional: rich Excel formatting ───────────────────────────────────────────
try:
    from openpyxl import load_workbook
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    EXCEL_FORMAT = True
except ImportError:
    EXCEL_FORMAT = False
    print("  [INFO] openpyxl not found — plain Excel only. Run: pip install openpyxl")

# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 3 CONFIG
# ══════════════════════════════════════════════════════════════════════════════
TRAIN_START      = dt.datetime(2016, 1, 1)
TRAIN_END        = dt.datetime(2021, 1, 1)
TEST_END         = dt.datetime.now()
MIN_ROWS         = 400        # minimum rows to consider an asset
PERIODS_PER_YEAR = 252

# Scoring thresholds for tier classification
TIER_S = {"sharpe": 0.8,  "ret": 30,  "max_dd": -20}   # Excellent
TIER_A = {"sharpe": 0.4,  "ret": 10,  "max_dd": -35}   # Good
TIER_B = {"sharpe": 0.1,  "ret": -10, "max_dd": -50}   # Decent
# Below TIER_B thresholds → Skip

# Output directory
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "screener_output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ══════════════════════════════════════════════════════════════════════════════
#  STOCK UNIVERSES  (Phase 3 — expanded, all 3 markets)
# ══════════════════════════════════════════════════════════════════════════════

# 🇺🇸 US — Large-cap + sector representation + ETFs
US_UNIVERSE = [
    # Tech mega-cap
    "AAPL", "MSFT", "NVDA", "GOOGL", "META", "AMZN", "TSLA", "AMD",
    # Financials
    "JPM", "BAC", "GS", "MS", "V", "MA", "BRK-B",
    # Healthcare
    "JNJ", "UNH", "LLY", "PFE", "ABBV", "MRK",
    # Energy
    "XOM", "CVX", "COP",
    # Consumer
    "PG", "KO", "PEP", "WMT", "COST", "MCD",
    # Industrials
    "CAT", "DE", "BA", "HON", "RTX",
    # ETFs (benchmark)
    "SPY", "QQQ", "GLD", "TLT", "IWM",
]

# 🇦🇺 ASX — All major sectors (ASX 200 coverage)
ASX_UNIVERSE = [
    # Big 4 Banks
    "CBA.AX", "WBC.AX", "ANZ.AX", "NAB.AX",
    # Resources — Iron Ore
    "BHP.AX", "RIO.AX", "FMG.AX", "S32.AX",
    # Energy
    "WDS.AX", "STO.AX", "BPT.AX",
    # Healthcare
    "CSL.AX", "RMD.AX", "COH.AX", "SHL.AX",
    # Retail / Consumer
    "WES.AX", "WOW.AX", "COL.AX", "JBH.AX",
    # Tech
    "XRO.AX", "WTC.AX", "ALU.AX",
    # Infrastructure / Utilities
    "TCL.AX", "APA.AX", "SKI.AX",
    # REITs
    "GMG.AX", "SCG.AX", "GPT.AX",
    # ETFs (benchmark)
    "IOZ.AX", "STW.AX", "VAS.AX",
]

# 🇯🇵 Japan (JPX/TSE) — Nikkei 225 major components across all sectors
JPX_UNIVERSE = [
    # Automotive
    "7203.T",   # Toyota Motor
    "7267.T",   # Honda Motor
    "7261.T",   # Mazda Motor
    "7272.T",   # Yamaha Motor
    "7269.T",   # Suzuki Motor
    # Electronics / Technology
    "6758.T",   # Sony Group
    "6501.T",   # Hitachi
    "6954.T",   # Fanuc
    "6902.T",   # Denso
    "6861.T",   # Keyence
    "6762.T",   # TDK
    "8035.T",   # Tokyo Electron
    "6857.T",   # Advantest
    "6723.T",   # Renesas Electronics
    "6594.T",   # Nidec (Nikkei component)
    # Semiconductors / Display
    "4063.T",   # Shin-Etsu Chemical
    "4523.T",   # Eisai
    # Telecom
    "9432.T",   # NTT
    "9433.T",   # KDDI
    "9434.T",   # SoftBank Corp
    # Internet / Media
    "9984.T",   # SoftBank Group
    "4689.T",   # Z Holdings (Yahoo Japan)
    "6098.T",   # Recruit Holdings
    "4385.T",   # Mercari
    # Gaming / Entertainment
    "7974.T",   # Nintendo
    "9684.T",   # Square Enix
    "7832.T",   # Bandai Namco
    # Financials — Banks
    "8306.T",   # MUFG (Mitsubishi UFJ)
    "8316.T",   # Sumitomo Mitsui
    "8411.T",   # Mizuho Financial
    # Financials — Insurance / Other
    "8750.T",   # Dai-ichi Life
    "8725.T",   # MS&AD Insurance
    # Consumer / Retail
    "3382.T",   # Seven & I Holdings
    "8267.T",   # Aeon
    "4661.T",   # Oriental Land (Disney Japan)
    "2914.T",   # Japan Tobacco
    # Industrials / Machinery
    "6301.T",   # Komatsu
    "6326.T",   # Kubota
    "7011.T",   # Mitsubishi Heavy Industries
    # Pharma / Healthcare
    "4502.T",   # Takeda Pharmaceutical
    "4519.T",   # Chugai Pharmaceutical
    # Trading Companies
    "8001.T",   # Itochu
    "8002.T",   # Marubeni
    "8058.T",   # Mitsubishi Corp
    # Transport / Infra
    "9022.T",   # Central Japan Railway (JR Tokai)
    "9020.T",   # East Japan Railway
    # ETFs (benchmark)
    "1321.T",   # Nikkei 225 ETF (Nomura)
    "1306.T",   # TOPIX ETF (Nomura)
]

UNIVERSE_MAP = {
    "US":  US_UNIVERSE,
    "ASX": ASX_UNIVERSE,
    "JPX": JPX_UNIVERSE,
}


# ══════════════════════════════════════════════════════════════════════════════
#  METRICS CALCULATOR
# ══════════════════════════════════════════════════════════════════════════════
def compute_metrics(price: pd.Series, position: pd.Series) -> dict:
    df = pd.DataFrame({"price": price})
    df["pos"]       = position.reindex(df.index).fillna(0)
    df["log_ret"]   = np.log(df["price"] / df["price"].shift(1))
    df["strat_ret"] = df["pos"] * df["log_ret"]
    df["cum"]       = np.exp(df["strat_ret"].cumsum())

    lr = df["strat_ret"].dropna()
    if len(lr) < 10:
        return {"Sharpe": np.nan, "TotalRet": np.nan, "MaxDD": np.nan,
                "WinRate": np.nan}

    ann    = np.exp(lr.mean() * PERIODS_PER_YEAR) - 1
    vol    = lr.std() * np.sqrt(PERIODS_PER_YEAR)
    sharpe = ann / vol if vol != 0 else np.nan
    cum    = df["cum"].dropna()
    mdd    = ((cum - cum.cummax()) / cum.cummax()).min() if len(cum) > 0 else np.nan
    total  = cum.iloc[-1] - 1 if len(cum) > 0 else np.nan
    wins   = (lr > 0).sum() / max(len(lr), 1)

    return {
        "Sharpe":   round(float(sharpe), 3) if not np.isnan(sharpe) else np.nan,
        "TotalRet": round(float(total) * 100, 2) if not np.isnan(total) else np.nan,
        "MaxDD":    round(float(mdd) * 100, 2) if not np.isnan(mdd) else np.nan,
        "WinRate":  round(float(wins) * 100, 1) if not np.isnan(wins) else np.nan,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  STRATEGY FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════
def strategy_ma(p, short=20, long=50):
    return (p.rolling(short).mean() > p.rolling(long).mean()).astype(int).shift(1)

def strategy_rsi(p, period=14, oversold=30, overbought=70):
    d   = p.diff()
    rsi = 100 - (100 / (1 + d.clip(lower=0).rolling(period).mean() /
                            (-d.clip(upper=0)).rolling(period).mean()))
    sig  = np.zeros(len(p)); hold = False
    for i, r in enumerate(rsi):
        if np.isnan(r): sig[i] = 0
        elif not hold and r < oversold:  hold = True;  sig[i] = 1
        elif hold and r > overbought:    hold = False; sig[i] = 0
        else:                            sig[i] = 1 if hold else 0
    return pd.Series(sig, index=p.index).shift(1)

def strategy_bb(p, window=20, num_std=2.0):
    mid = p.rolling(window).mean()
    low = mid - num_std * p.rolling(window).std()
    sig = np.zeros(len(p)); hold = False
    for i in range(len(p)):
        pi, m, l = p.iloc[i], mid.iloc[i], low.iloc[i]
        if np.isnan(l): sig[i] = 0
        elif not hold and pi <= l: hold = True;  sig[i] = 1
        elif hold and pi >= m:     hold = False; sig[i] = 0
        else:                      sig[i] = 1 if hold else 0
    return pd.Series(sig, index=p.index).shift(1)

def strategy_macd(p, fast=12, slow=26, signal_p=9):
    macd = p.ewm(span=fast).mean() - p.ewm(span=slow).mean()
    return (macd > macd.ewm(span=signal_p).mean()).astype(int).shift(1)

def strategy_mr(p, window=20, threshold=1.5):
    z   = (p - p.rolling(window).mean()) / p.rolling(window).std()
    sig = np.zeros(len(p)); hold = False
    for i, zi in enumerate(z):
        if np.isnan(zi): sig[i] = 0
        elif not hold and zi < -threshold: hold = True;  sig[i] = 1
        elif hold and zi >= 0:             hold = False; sig[i] = 0
        else:                              sig[i] = 1 if hold else 0
    return pd.Series(sig, index=p.index).shift(1)

STRATEGY_FNS = {
    "MA Crossover":    strategy_ma,
    "RSI":             strategy_rsi,
    "Bollinger Bands": strategy_bb,
    "MACD":            strategy_macd,
    "Mean Reversion":  strategy_mr,
}

STRATEGY_GRIDS = {
    "MA Crossover":    [{"short": s, "long": l}
                        for s in [10, 20, 50] for l in [50, 100, 200] if s < l],
    "RSI":             [{"period": p, "oversold": os, "overbought": ob}
                        for p in [10, 14, 21] for os, ob in [(25,65),(30,70),(35,75)]],
    "Bollinger Bands": [{"window": w, "num_std": s}
                        for w in [10, 20, 30] for s in [1.5, 2.0, 2.5]],
    "MACD":            [{"fast": f, "slow": s, "signal_p": sg}
                        for f in [8, 12] for s in [21, 26] for sg in [7, 9] if f < s],
    "Mean Reversion":  [{"window": w, "threshold": t}
                        for w in [10, 20, 40] for t in [1.0, 1.5, 2.0]],
}


# ══════════════════════════════════════════════════════════════════════════════
#  SCORING & TIERING
# ══════════════════════════════════════════════════════════════════════════════
def score_asset(row: dict):
    """
    Composite score (0–100) combining:
    - OUT Sharpe         (40 pts max) — quality of risk-adjusted return
    - OUT Total Return   (25 pts max) — raw performance
    - Max DD Protection  (20 pts max) — capital preservation
    - DD Saved vs B&H    (15 pts max) — strategy added value over hold

    Returns (tier, score)
    """
    sharpe  = row.get("OUT Sharpe", np.nan)
    ret     = row.get("OUT Strat Ret %", np.nan)
    mdd     = row.get("OUT Strat Max DD %", np.nan)
    dd_save = row.get("DD Saved %", np.nan)

    if any(pd.isna(v) for v in [sharpe, ret, mdd]):
        return ("Skip", 0.0)

    # Sharpe: 0.0 → 0 pts,  1.5+ → 40 pts
    s_pts  = min(max(sharpe / 1.5, 0), 1) * 40
    # Return: -30% → 0 pts, 100%+ → 25 pts  (log-scaled)
    r_norm = max(ret + 30, 0) / 130
    r_pts  = min(r_norm, 1) * 25
    # MaxDD: -60% → 0 pts,  0% → 20 pts
    d_pts  = min(max((60 + mdd) / 60, 0), 1) * 20
    # DD Saved: -20% → 0 pts,  40%+ → 15 pts
    if not pd.isna(dd_save):
        ds_pts = min(max((dd_save + 20) / 60, 0), 1) * 15
    else:
        ds_pts = 0

    score = round(s_pts + r_pts + d_pts + ds_pts, 1)

    if sharpe >= TIER_S["sharpe"] and ret >= TIER_S["ret"] and mdd >= TIER_S["max_dd"]:
        tier = "S"
    elif sharpe >= TIER_A["sharpe"] and ret >= TIER_A["ret"] and mdd >= TIER_A["max_dd"]:
        tier = "A"
    elif sharpe >= TIER_B["sharpe"] and mdd >= TIER_B["max_dd"]:
        tier = "B"
    else:
        tier = "Skip"

    return (tier, score)


# ══════════════════════════════════════════════════════════════════════════════
#  MARKET SCREENER — single market pass
# ══════════════════════════════════════════════════════════════════════════════
def screen_market(market_name: str, assets: list, price_data: dict) -> pd.DataFrame:
    today = dt.datetime.now().strftime("%Y-%m-%d")
    results = []

    print(f"\n  ── Screening {market_name} ({len([a for a in assets if a in price_data])} valid assets) ──")

    for asset in assets:
        if asset not in price_data:
            continue

        full  = price_data[asset]
        train = full[full.index < TRAIN_END]
        test  = full[full.index >= TRAIN_END]

        if len(train) < 100 or len(test) < 50:
            continue

        # ── Best strategy search (train period) ──────────────────────────────
        best_sharpe = -999
        best_strat  = None
        best_params = None

        for strat_name, fn in STRATEGY_FNS.items():
            for params in STRATEGY_GRIDS[strat_name]:
                try:
                    pos = fn(train, **params)
                    m   = compute_metrics(train, pos)
                    if not np.isnan(m["Sharpe"]) and m["Sharpe"] > best_sharpe:
                        best_sharpe = m["Sharpe"]
                        best_strat  = strat_name
                        best_params = params
                except Exception:
                    continue

        if best_strat is None:
            continue

        # ── Out-of-sample (test period) ──────────────────────────────────────
        pos_test = STRATEGY_FNS[best_strat](test, **best_params)
        tm       = compute_metrics(test, pos_test)

        # ── B&H benchmark ────────────────────────────────────────────────────
        log      = np.log(test / test.shift(1)).dropna()
        bah_ret  = (np.exp(log.cumsum()).iloc[-1] - 1) * 100 if len(log) > 0 else np.nan
        bah_cum  = np.exp(log.cumsum()) if len(log) > 0 else pd.Series(dtype=float)
        bah_mdd  = ((bah_cum - bah_cum.cummax()) / bah_cum.cummax()).min() * 100 \
                   if len(bah_cum) > 0 else np.nan
        dd_saved = round(tm["MaxDD"] - float(bah_mdd), 2) \
                   if not np.isnan(bah_mdd) else np.nan

        row = {
            "Market":             market_name,
            "Asset":              asset,
            "Best Strategy":      best_strat,
            "Best Params":        str(best_params),
            "Train Sharpe":       round(best_sharpe, 3),
            "OUT Sharpe":         tm["Sharpe"],
            "OUT Win Rate %":     tm["WinRate"],
            "OUT Strat Ret %":    tm["TotalRet"],
            "OUT B&H Ret %":      round(float(bah_ret), 2) if not np.isnan(bah_ret) else np.nan,
            "OUT Strat Max DD %": tm["MaxDD"],
            "OUT B&H Max DD %":   round(float(bah_mdd), 2) if not np.isnan(bah_mdd) else np.nan,
            "DD Saved %":         dd_saved,
            "Beats B&H":          (tm["TotalRet"] or 0) > (bah_ret or 0),
            "Run Date":           today,
        }

        tier, score = score_asset(row)
        row["Tier"]  = tier
        row["Score"] = score
        results.append(row)

        # Console output
        tier_icon = {"S": "⭐", "A": "✅", "B": "🔵", "Skip": "⬜"}.get(tier, "")
        beats_icon = "✓" if row["Beats B&H"] else "✗"
        dd_str     = f"{dd_saved:+.1f}%" if not np.isnan(dd_saved) else "N/A"
        print(f"    {tier_icon} [{tier}] {asset:14} → {best_strat:18} | "
              f"Sharpe: {tm['Sharpe']:5.2f} | "
              f"Ret: {(tm['TotalRet'] or 0):6.0f}% | "
              f"DD: {(tm['MaxDD'] or 0):5.1f}% | "
              f"DDsaved: {dd_str:>8} | B&H {beats_icon} | Score: {score:.0f}/100")

    return pd.DataFrame(results)


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL REPORTER
# ══════════════════════════════════════════════════════════════════════════════
TIER_COLORS = {
    "S":    "FFD700",   # Gold
    "A":    "90EE90",   # Light green
    "B":    "87CEEB",   # Sky blue
    "Skip": "D3D3D3",   # Light grey
}

STRAT_COLORS = {
    "MA Crossover":    "AED6F1",
    "RSI":             "FAD7A0",
    "Bollinger Bands": "A9DFBF",
    "MACD":            "D7BDE2",
    "Mean Reversion":  "F1948A",
}

def apply_sheet_formatting(ws, df):
    if not EXCEL_FORMAT:
        return
    try:
        THIN   = Side(style="thin", color="CCCCCC")
        BRD    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        HDR_F  = PatternFill("solid", fgColor="1C2833")
        GRN    = PatternFill("solid", fgColor="D5F5E3")
        RED    = PatternFill("solid", fgColor="FADBD8")

        headers = [c.value for c in ws[1]]
        col_idx = {h: i+1 for i, h in enumerate(headers)}

        # Header row
        for cell in ws[1]:
            cell.fill      = HDR_F
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = BRD
        ws.row_dimensions[1].height = 30

        # Data rows
        for row in ws.iter_rows(min_row=2):
            strat_val = row[col_idx["Best Strategy"]-1].value if "Best Strategy" in col_idx else None
            tier_val  = row[col_idx["Tier"]-1].value          if "Tier"          in col_idx else None
            beats_val = row[col_idx["Beats B&H"]-1].value     if "Beats B&H"     in col_idx else None

            for cell in row:
                cell.border    = BRD
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font      = Font(size=10)
                h = headers[cell.column - 1]

                if h == "Best Strategy" and strat_val:
                    cell.fill = PatternFill("solid", fgColor=STRAT_COLORS.get(strat_val, "FFFFFF"))
                    cell.font = Font(bold=True, size=10)
                elif h == "Tier" and tier_val:
                    cell.fill = PatternFill("solid", fgColor=TIER_COLORS.get(tier_val, "FFFFFF"))
                    cell.font = Font(bold=True, size=10)
                elif h == "Score":
                    try:
                        v = float(cell.value or 0)
                        if v >= 70:   cell.fill = PatternFill("solid", fgColor="27AE60")
                        elif v >= 50: cell.fill = PatternFill("solid", fgColor="F39C12")
                        elif v >= 30: cell.fill = PatternFill("solid", fgColor="AED6F1")
                        else:         cell.fill = PatternFill("solid", fgColor="E8E8E8")
                    except: pass
                elif h == "OUT Sharpe":
                    try:
                        v = float(cell.value or 0)
                        if v >= 0.8:  cell.fill = GRN
                        elif v < 0:   cell.fill = RED
                    except: pass
                elif h == "OUT Strat Ret %":
                    try:
                        v = float(cell.value or 0)
                        if v >= 30:   cell.fill = GRN
                        elif v < -10: cell.fill = RED
                    except: pass
                elif h in ("OUT Strat Max DD %", "OUT B&H Max DD %"):
                    try:
                        v = float(cell.value or 0)
                        if v >= -20:  cell.fill = GRN
                        elif v < -40: cell.fill = RED
                    except: pass
                elif h == "DD Saved %":
                    try:
                        v = float(cell.value or 0)
                        if v >= 10:  cell.fill = GRN
                        elif v < 0:  cell.fill = RED
                    except: pass
                elif h == "Beats B&H":
                    cell.fill = GRN if beats_val else RED

        # Column widths
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 32)

        # Freeze header
        ws.freeze_panes = "A2"

    except Exception as e:
        print(f"  [WARN] Formatting failed: {e}")


def write_excel_report(all_results: dict, today: str) -> str:
    """Write multi-tab Excel report with per-market tabs + summary."""
    fname = f"Goofy_Phase3_{today}.xlsx"
    path  = os.path.join(OUTPUT_DIR, fname)

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    all_df_list = []
    market_flags = {"US": "🇺🇸", "ASX": "🇦🇺", "JPX": "🇯🇵"}

    # ── Per-market tabs ───────────────────────────────────────────────────────
    for market_name, df in all_results.items():
        if df.empty:
            continue
        df_sorted = df.sort_values("Score", ascending=False).reset_index(drop=True)
        flag = market_flags.get(market_name, "")
        ws = wb.create_sheet(title=f"{flag} {market_name}")

        cols_to_show = [c for c in df_sorted.columns if c != "Best Params"]
        ws.append(cols_to_show)
        for _, row in df_sorted[cols_to_show].iterrows():
            ws.append([row[c] for c in cols_to_show])

        apply_sheet_formatting(ws, df_sorted[cols_to_show])
        all_df_list.append(df_sorted)

    # ── Top Performers tab ────────────────────────────────────────────────────
    if all_df_list:
        combined   = pd.concat(all_df_list, ignore_index=True)
        top_picks  = combined[combined["Tier"].isin(["S", "A"])].sort_values(
                        "Score", ascending=False).reset_index(drop=True)
        decent     = combined[combined["Tier"] == "B"].sort_values(
                        "Score", ascending=False).reset_index(drop=True)

        ws_top = wb.create_sheet(title="⭐ Top Performers", index=0)
        ws_top.append(["GOOFY SCREENER — PHASE 3 RESULTS"])
        ws_top.append([f"Run: {today}  |  Markets: US + ASX + JPX  |  "
                       f"Assets screened: {len(combined)}"])
        ws_top.append([])

        # Tier legend
        ws_top.append(["TIER", "CRITERIA"])
        ws_top.append(["S (⭐ Excellent)", "Sharpe ≥ 0.8, Return ≥ 30%, Max DD ≥ -20%"])
        ws_top.append(["A (✅ Good)",       "Sharpe ≥ 0.4, Return ≥ 10%, Max DD ≥ -35%"])
        ws_top.append(["B (🔵 Decent)",    "Sharpe ≥ 0.1, Max DD ≥ -50%"])
        ws_top.append(["Skip (⬜)",         "Below all thresholds"])
        ws_top.append([])

        # Summary counts
        ws_top.append(["TIER BREAKDOWN"])
        for tier in ["S", "A", "B", "Skip"]:
            count = len(combined[combined["Tier"] == tier])
            ws_top.append([tier, count])
        ws_top.append([])

        ws_top.append(["── S & A TIER (Best Opportunities) ──"])
        if not top_picks.empty:
            cols_show = ["Market", "Asset", "Best Strategy", "Tier", "Score",
                         "OUT Sharpe", "OUT Strat Ret %", "OUT B&H Ret %",
                         "OUT Strat Max DD %", "DD Saved %", "Beats B&H"]
            ws_top.append(cols_show)
            for _, row in top_picks[cols_show].iterrows():
                ws_top.append([row[c] for c in cols_show])
        else:
            ws_top.append(["No S or A tier assets found this run."])

        ws_top.append([])
        ws_top.append(["── B TIER (Decent — worth watching) ──"])
        if not decent.empty:
            cols_show = ["Market", "Asset", "Best Strategy", "Score",
                         "OUT Sharpe", "OUT Strat Ret %", "OUT Strat Max DD %", "DD Saved %"]
            ws_top.append(cols_show)
            for _, row in decent[cols_show].iterrows():
                ws_top.append([row[c] for c in cols_show])
        else:
            ws_top.append(["No B tier assets found."])

        # Style the summary sheet header
        if EXCEL_FORMAT:
            try:
                ws_top["A1"].font = Font(bold=True, size=14, color="1C2833")
                ws_top["A2"].font = Font(italic=True, size=10, color="555555")
            except: pass

    # ── Strategy Distribution tab ─────────────────────────────────────────────
    if all_df_list:
        combined_all = pd.concat(all_df_list, ignore_index=True)
        ws_dist = wb.create_sheet(title="📊 Strategy Distribution")

        ws_dist.append(["Strategy", "US Count", "ASX Count", "JPX Count", "Total"])
        for strat in STRATEGY_FNS.keys():
            row_data = [strat]
            total = 0
            for market in ["US", "ASX", "JPX"]:
                if market in all_results and not all_results[market].empty:
                    c = (all_results[market]["Best Strategy"] == strat).sum()
                else:
                    c = 0
                row_data.append(c)
                total += c
            row_data.append(total)
            ws_dist.append(row_data)

        ws_dist.append([])
        ws_dist.append(["Asset-Strategy Fit (from Phase 1-2 research):"])
        ws_dist.append(["NVDA/TSLA", "→ MACD (momentum)"])
        ws_dist.append(["Sony (6758.T)", "→ Bollinger Bands (mean-reverting)"])
        ws_dist.append(["SPY/ETFs", "→ MA Crossover (efficient markets)"])
        ws_dist.append(["Banks (CBA, MUFG)", "→ RSI / BB"])

        if EXCEL_FORMAT:
            try:
                for cell in ws_dist[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="1C2833")
            except: pass

    wb.save(path)
    return path


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="Goofy Phase 3 Multi-Market Screener")
    parser.add_argument("--market", choices=["US", "ASX", "JPX", "ALL"],
                        default="ALL", help="Which market to screen (default: ALL)")
    args = parser.parse_args()

    today = dt.datetime.now().strftime("%Y-%m-%d")

    markets_to_run = (["US", "ASX", "JPX"] if args.market == "ALL"
                      else [args.market])

    print(f"""
╔══════════════════════════════════════════════════════════════════════╗
║  GOOFY SCREENER — PHASE 3  |  {today}                    ║
║  Markets: {', '.join(markets_to_run):52} ║
║  Train: 2016-2021  |  Test: 2021-{today[:4]}                        ║
╚══════════════════════════════════════════════════════════════════════╝""")

    # ── Step 1: Download all data ─────────────────────────────────────────────
    all_assets = []
    for m in markets_to_run:
        all_assets.extend(UNIVERSE_MAP[m])
    # deduplicate
    seen = set(); unique_assets = []
    for a in all_assets:
        if a not in seen: unique_assets.append(a); seen.add(a)

    print(f"\n[1/3] Downloading data for {len(unique_assets)} assets...\n")
    price_data = {}
    for asset in unique_assets:
        try:
            raw = yf.download(asset, start=TRAIN_START, end=TEST_END,
                              auto_adjust=True, progress=False)
            if not raw.empty and len(raw) >= MIN_ROWS:
                close = raw["Close"].squeeze()
                if isinstance(close, pd.DataFrame):
                    close = close.iloc[:, 0]
                price_data[asset] = close
                print(f"  ✓ {asset:14} | {len(raw)} rows")
            else:
                print(f"  ✗ {asset:14} | skipped (only {len(raw)} rows)")
        except Exception as e:
            print(f"  ✗ {asset:14} | error: {e}")

    print(f"\n  → {len(price_data)} assets ready for screening.\n")

    # ── Step 2: Run screener per market ───────────────────────────────────────
    print(f"[2/3] Running strategy screening...\n")
    all_results = {}
    for m in markets_to_run:
        all_results[m] = screen_market(m, UNIVERSE_MAP[m], price_data)

    # ── Step 3: Save & summarise ──────────────────────────────────────────────
    print(f"\n[3/3] Saving report...\n")

    all_dfs = [df for df in all_results.values() if not df.empty]
    if not all_dfs:
        print("  No results to save.")
        return

    combined = pd.concat(all_dfs, ignore_index=True)

    # ── Console Summary ────────────────────────────────────────────────────────
    print(f"\n{'═'*70}")
    print(f"  GOOFY SCREENER — PHASE 3 SUMMARY  |  {today}")
    print(f"{'═'*70}")
    print(f"  Assets screened:  {len(combined)}")
    for market in markets_to_run:
        df_m = all_results.get(market, pd.DataFrame())
        if not df_m.empty:
            print(f"    {market:5}: {len(df_m)} assets")

    print(f"\n  ── Tier Breakdown ──")
    for tier in ["S", "A", "B", "Skip"]:
        icon  = {"S":"⭐","A":"✅","B":"🔵","Skip":"⬜"}.get(tier,"")
        count = len(combined[combined["Tier"] == tier])
        bar   = "█" * count
        print(f"    {icon} {tier:5}: {count:3}  {bar}")

    print(f"\n  ── ⭐ S-Tier (Excellent) ──")
    s_tier = combined[combined["Tier"] == "S"].sort_values("Score", ascending=False)
    if not s_tier.empty:
        for _, r in s_tier.iterrows():
            print(f"    {r['Market']:5} {r['Asset']:14} | {r['Best Strategy']:18} | "
                  f"Sharpe: {r['OUT Sharpe']:5.2f} | Ret: {r['OUT Strat Ret %']:6.1f}% | "
                  f"DD: {r['OUT Strat Max DD %']:5.1f}% | Score: {r['Score']:.0f}/100")
    else:
        print("    (none this run)")

    print(f"\n  ── ✅ A-Tier (Good) ──")
    a_tier = combined[combined["Tier"] == "A"].sort_values("Score", ascending=False)
    if not a_tier.empty:
        for _, r in a_tier.iterrows():
            print(f"    {r['Market']:5} {r['Asset']:14} | {r['Best Strategy']:18} | "
                  f"Sharpe: {r['OUT Sharpe']:5.2f} | Ret: {r['OUT Strat Ret %']:6.1f}% | "
                  f"DD: {r['OUT Strat Max DD %']:5.1f}% | Score: {r['Score']:.0f}/100")
    else:
        print("    (none this run)")

    print(f"\n  ── 🔵 B-Tier (Decent — worth watching) ──")
    b_tier = combined[combined["Tier"] == "B"].sort_values("Score", ascending=False).head(10)
    if not b_tier.empty:
        for _, r in b_tier.iterrows():
            print(f"    {r['Market']:5} {r['Asset']:14} | {r['Best Strategy']:18} | "
                  f"Sharpe: {r['OUT Sharpe']:5.2f} | Ret: {r['OUT Strat Ret %']:6.1f}% | Score: {r['Score']:.0f}/100")

    print(f"\n  ── Strategy Distribution ──")
    for strat, count in combined["Best Strategy"].value_counts().items():
        pct = count / len(combined) * 100
        bar = "█" * int(pct / 5)
        print(f"    {strat:20}: {count:3}  ({pct:.0f}%)  {bar}")

    print(f"\n  ── Top 5 by Composite Score ──")
    top5 = combined.nlargest(5, "Score")[
        ["Market","Asset","Best Strategy","Tier","Score","OUT Sharpe","OUT Strat Ret %","OUT Strat Max DD %"]
    ]
    print(top5.to_string(index=False))

    print(f"\n  ── Top 5 DD Protectors (DD Saved vs B&H) ──")
    top_dd = combined.nlargest(5, "DD Saved %")[
        ["Market","Asset","Best Strategy","OUT Strat Max DD %","OUT B&H Max DD %","DD Saved %"]
    ]
    print(top_dd.to_string(index=False))

    # ── Save Excel ─────────────────────────────────────────────────────────────
    xlsx_path = write_excel_report(all_results, today)
    print(f"\n  Report saved → {os.path.basename(xlsx_path)}")
    print(f"  Full path:  {xlsx_path}")
    print(f"\n{'═'*70}\n")


if __name__ == "__main__":
    main()
