"""
╔══════════════════════════════════════════════════════════════════════╗
║           GOOFY SCREENER — PHASE 4                                   ║
║           Multi-Market Screener  +  Regime-Aware Verdict Layer       ║
║                                                                      ║
║   Inherits everything from Phase 3 (US/ASX/JPX, 5 strategies, S/A/B  ║
║   tiering, Excel report) — adds a runtime "should I trade today?"   ║
║   verdict on top of the picked strategy, using empirical asset-      ║
║   specific gates from `asset_specific_gates.json`.                   ║
║                                                                      ║
║   The historical backtest is UNCHANGED — that preserves your         ║
║   research truth. Gating is applied AFTER the screener picks a       ║
║   strategy, so it's a pure execution-layer decision (the same way    ║
║   real funds split alpha research from execution).                   ║
║                                                                      ║
║   Three new columns:                                                 ║
║       Current Trend     —  Bull / Sideways / Bear  (today)           ║
║       Current Vol       —  Low / Normal / High      (today)          ║
║       Today's Verdict   —  TRADE / STAND DOWN       (today)          ║
║                                                                      ║
║   Run:  python goofy_screener_phase4.py                              ║
║   Args: --market US        (default: ALL)                            ║
║         --market ASX                                                 ║
║         --market JPX                                                 ║
║         --market ALL                                                 ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import yfinance as yf
import numpy as np
import pandas as pd
import datetime as dt
import os, sys, warnings, argparse
warnings.filterwarnings("ignore")

# Phase 4 add-on: regime detection + asset-specific gates
from regime_detector import (
    label_regimes,
    get_allowed_regimes,
    load_asset_gates,
    ASSET_SPECIFIC_GATES,
)

# ── Optional: rich Excel formatting ───────────────────────────────────────────
try:
    from openpyxl import load_workbook
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    EXCEL_FORMAT = True
except ImportError:
    EXCEL_FORMAT = False
    print("  [INFO] openpyxl not found — plain Excel only. Run: pip install openpyxl")

# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 4 CONFIG
# ══════════════════════════════════════════════════════════════════════════════
TRAIN_START      = dt.datetime(2016, 1, 1)
TRAIN_END        = dt.datetime(2021, 1, 1)
TEST_END         = dt.datetime.now()
MIN_ROWS         = 400        # minimum rows to consider an asset
PERIODS_PER_YEAR = 252

# Scoring thresholds for tier classification
TIER_S = {"sharpe": 0.8,  "ret": 30,  "max_dd": -20}   # Excellent
TIER_A = {"sharpe": 0.4,  "ret": 10,  "max_dd": -35}   # Good
TIER_B = {"sharpe": 0.1,  "ret": -10, "max_dd": -50}   # Decent
# Below TIER_B thresholds → Skip

# Output directory
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "screener_output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Long-name (used by the screener) → short-name (used by the gate dict).
# regime_detector uses short names like "MA", "BB", "MeanReversion".
LONG_TO_SHORT = {
    "MA Crossover":    "MA",
    "RSI":             "RSI",
    "Bollinger Bands": "BB",
    "MACD":            "MACD",
    "Mean Reversion":  "MeanReversion",
}

# ══════════════════════════════════════════════════════════════════════════════
#  STOCK UNIVERSES  (same as Phase 3)
# ══════════════════════════════════════════════════════════════════════════════

# 🇺🇸 US — Large-cap + sector representation + ETFs
US_UNIVERSE = [
    # Tech mega-cap
    "AAPL", "MSFT", "NVDA", "GOOGL", "META", "AMZN", "TSLA", "AMD",
    # Financials
    "JPM", "BAC", "GS", "MS", "V", "MA", "BRK-B",
    # Healthcare
    "JNJ", "UNH", "LLY", "PFE", "ABBV", "MRK",
    # Energy
    "XOM", "CVX", "COP",
    # Consumer
    "PG", "KO", "PEP", "WMT", "COST", "MCD",
    # Industrials
    "CAT", "DE", "BA", "HON", "RTX",
    # ETFs (benchmark)
    "SPY", "QQQ", "GLD", "TLT", "IWM",
]

# 🇦🇺 ASX — All major sectors (ASX 200 coverage)
ASX_UNIVERSE = [
    # Big 4 Banks
    "CBA.AX", "WBC.AX", "ANZ.AX", "NAB.AX",
    # Resources — Iron Ore
    "BHP.AX", "RIO.AX", "FMG.AX", "S32.AX",
    # Energy
    "WDS.AX", "STO.AX", "BPT.AX",
    # Healthcare
    "CSL.AX", "RMD.AX", "COH.AX", "SHL.AX",
    # Retail / Consumer
    "WES.AX", "WOW.AX", "COL.AX", "JBH.AX",
    # Tech
    "XRO.AX", "WTC.AX", "ALU.AX",
    # Infrastructure / Utilities
    "TCL.AX", "APA.AX", "SKI.AX",
    # REITs
    "GMG.AX", "SCG.AX", "GPT.AX",
    # ETFs (benchmark)
    "IOZ.AX", "STW.AX", "VAS.AX",
]

# 🇯🇵 Japan (JPX/TSE) — Nikkei 225 major components across all sectors
JPX_UNIVERSE = [
    # Automotive
    "7203.T",   # Toyota Motor
    "7267.T",   # Honda Motor
    "7261.T",   # Mazda Motor
    "7272.T",   # Yamaha Motor
    "7269.T",   # Suzuki Motor
    # Electronics / Technology
    "6758.T",   # Sony Group
    "6501.T",   # Hitachi
    "6954.T",   # Fanuc
    "6902.T",   # Denso
    "6861.T",   # Keyence
    "6762.T",   # TDK
    "8035.T",   # Tokyo Electron
    "6857.T",   # Advantest
    "6723.T",   # Renesas Electronics
    "6594.T",   # Nidec (Nikkei component)
    # Semiconductors / Display
    "4063.T",   # Shin-Etsu Chemical
    "4523.T",   # Eisai
    # Telecom
    "9432.T",   # NTT
    "9433.T",   # KDDI
    "9434.T",   # SoftBank Corp
    # Internet / Media
    "9984.T",   # SoftBank Group
    "4689.T",   # Z Holdings (Yahoo Japan)
    "6098.T",   # Recruit Holdings
    "4385.T",   # Mercari
    # Gaming / Entertainment
    "7974.T",   # Nintendo
    "9684.T",   # Square Enix
    "7832.T",   # Bandai Namco
    # Financials — Banks
    "8306.T",   # MUFG (Mitsubishi UFJ)
    "8316.T",   # Sumitomo Mitsui
    "8411.T",   # Mizuho Financial
    # Financials — Insurance / Other
    "8750.T",   # Dai-ichi Life
    "8725.T",   # MS&AD Insurance
    # Consumer / Retail
    "3382.T",   # Seven & I Holdings
    "8267.T",   # Aeon
    "4661.T",   # Oriental Land (Disney Japan)
    "2914.T",   # Japan Tobacco
    # Industrials / Machinery
    "6301.T",   # Komatsu
    "6326.T",   # Kubota
    "7011.T",   # Mitsubishi Heavy Industries
    # Pharma / Healthcare
    "4502.T",   # Takeda Pharmaceutical
    "4519.T",   # Chugai Pharmaceutical
    # Trading Companies
    "8001.T",   # Itochu
    "8002.T",   # Marubeni
    "8058.T",   # Mitsubishi Corp
    # Transport / Infra
    "9022.T",   # Central Japan Railway (JR Tokai)
    "9020.T",   # East Japan Railway
    # ETFs (benchmark)
    "1321.T",   # Nikkei 225 ETF (Nomura)
    "1306.T",   # TOPIX ETF (Nomura)
]

UNIVERSE_MAP = {
    "US":  US_UNIVERSE,
    "ASX": ASX_UNIVERSE,
    "JPX": JPX_UNIVERSE,
}


# ══════════════════════════════════════════════════════════════════════════════
#  METRICS CALCULATOR
# ══════════════════════════════════════════════════════════════════════════════
def compute_metrics(price: pd.Series, position: pd.Series) -> dict:
    df = pd.DataFrame({"price": price})
    df["pos"]       = position.reindex(df.index).fillna(0)
    df["log_ret"]   = np.log(df["price"] / df["price"].shift(1))
    df["strat_ret"] = df["pos"] * df["log_ret"]
    df["cum"]       = np.exp(df["strat_ret"].cumsum())

    lr = df["strat_ret"].dropna()
    if len(lr) < 10:
        return {"Sharpe": np.nan, "TotalRet": np.nan, "MaxDD": np.nan,
                "WinRate": np.nan}

    ann    = np.exp(lr.mean() * PERIODS_PER_YEAR) - 1
    vol    = lr.std() * np.sqrt(PERIODS_PER_YEAR)
    sharpe = ann / vol if vol != 0 else np.nan
    cum    = df["cum"].dropna()
    mdd    = ((cum - cum.cummax()) / cum.cummax()).min() if len(cum) > 0 else np.nan
    total  = cum.iloc[-1] - 1 if len(cum) > 0 else np.nan
    wins   = (lr > 0).sum() / max(len(lr), 1)

    return {
        "Sharpe":   round(float(sharpe), 3) if not np.isnan(sharpe) else np.nan,
        "TotalRet": round(float(total) * 100, 2) if not np.isnan(total) else np.nan,
        "MaxDD":    round(float(mdd) * 100, 2) if not np.isnan(mdd) else np.nan,
        "WinRate":  round(float(wins) * 100, 1) if not np.isnan(wins) else np.nan,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  STRATEGY FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════
def strategy_ma(p, short=20, long=50):
    return (p.rolling(short).mean() > p.rolling(long).mean()).astype(int).shift(1)

def strategy_rsi(p, period=14, oversold=30, overbought=70):
    d   = p.diff()
    rsi = 100 - (100 / (1 + d.clip(lower=0).rolling(period).mean() /
                            (-d.clip(upper=0)).rolling(period).mean()))
    sig  = np.zeros(len(p)); hold = False
    for i, r in enumerate(rsi):
        if np.isnan(r): sig[i] = 0
        elif not hold and r < oversold:  hold = True;  sig[i] = 1
        elif hold and r > overbought:    hold = False; sig[i] = 0
        else:                            sig[i] = 1 if hold else 0
    return pd.Series(sig, index=p.index).shift(1)

def strategy_bb(p, window=20, num_std=2.0):
    mid = p.rolling(window).mean()
    low = mid - num_std * p.rolling(window).std()
    sig = np.zeros(len(p)); hold = False
    for i in range(len(p)):
        pi, m, l = p.iloc[i], mid.iloc[i], low.iloc[i]
        if np.isnan(l): sig[i] = 0
        elif not hold and pi <= l: hold = True;  sig[i] = 1
        elif hold and pi >= m:     hold = False; sig[i] = 0
        else:                      sig[i] = 1 if hold else 0
    return pd.Series(sig, index=p.index).shift(1)

def strategy_macd(p, fast=12, slow=26, signal_p=9):
    macd = p.ewm(span=fast).mean() - p.ewm(span=slow).mean()
    return (macd > macd.ewm(span=signal_p).mean()).astype(int).shift(1)

def strategy_mr(p, window=20, threshold=1.5):
    z   = (p - p.rolling(window).mean()) / p.rolling(window).std()
    sig = np.zeros(len(p)); hold = False
    for i, zi in enumerate(z):
        if np.isnan(zi): sig[i] = 0
        elif not hold and zi < -threshold: hold = True;  sig[i] = 1
        elif hold and zi >= 0:             hold = False; sig[i] = 0
        else:                              sig[i] = 1 if hold else 0
    return pd.Series(sig, index=p.index).shift(1)

STRATEGY_FNS = {
    "MA Crossover":    strategy_ma,
    "RSI":             strategy_rsi,
    "Bollinger Bands": strategy_bb,
    "MACD":            strategy_macd,
    "Mean Reversion":  strategy_mr,
}

STRATEGY_GRIDS = {
    "MA Crossover":    [{"short": s, "long": l}
                        for s in [10, 20, 50] for l in [50, 100, 200] if s < l],
    "RSI":             [{"period": p, "oversold": os, "overbought": ob}
                        for p in [10, 14, 21] for os, ob in [(25,65),(30,70),(35,75)]],
    "Bollinger Bands": [{"window": w, "num_std": s}
                        for w in [10, 20, 30] for s in [1.5, 2.0, 2.5]],
    "MACD":            [{"fast": f, "slow": s, "signal_p": sg}
                        for f in [8, 12] for s in [21, 26] for sg in [7, 9] if f < s],
    "Mean Reversion":  [{"window": w, "threshold": t}
                        for w in [10, 20, 40] for t in [1.0, 1.5, 2.0]],
}


# ══════════════════════════════════════════════════════════════════════════════
#  SCORING & TIERING
# ══════════════════════════════════════════════════════════════════════════════
def score_asset(row: dict):
    """
    Composite score (0–100) combining:
    - OUT Sharpe         (40 pts max) — quality of risk-adjusted return
    - OUT Total Return   (25 pts max) — raw performance
    - Max DD Protection  (20 pts max) — capital preservation
    - DD Saved vs B&H    (15 pts max) — strategy added value over hold

    Returns (tier, score)
    """
    sharpe  = row.get("OUT Sharpe", np.nan)
    ret     = row.get("OUT Strat Ret %", np.nan)
    mdd     = row.get("OUT Strat Max DD %", np.nan)
    dd_save = row.get("DD Saved %", np.nan)

    if any(pd.isna(v) for v in [sharpe, ret, mdd]):
        return ("Skip", 0.0)

    # Sharpe: 0.0 → 0 pts,  1.5+ → 40 pts
    s_pts  = min(max(sharpe / 1.5, 0), 1) * 40
    # Return: -30% → 0 pts, 100%+ → 25 pts  (log-scaled)
    r_norm = max(ret + 30, 0) / 130
    r_pts  = min(r_norm, 1) * 25
    # MaxDD: -60% → 0 pts,  0% → 20 pts
    d_pts  = min(max((60 + mdd) / 60, 0), 1) * 20
    # DD Saved: -20% → 0 pts,  40%+ → 15 pts
    if not pd.isna(dd_save):
        ds_pts = min(max((dd_save + 20) / 60, 0), 1) * 15
    else:
        ds_pts = 0

    score = round(s_pts + r_pts + d_pts + ds_pts, 1)

    if sharpe >= TIER_S["sharpe"] and ret >= TIER_S["ret"] and mdd >= TIER_S["max_dd"]:
        tier = "S"
    elif sharpe >= TIER_A["sharpe"] and ret >= TIER_A["ret"] and mdd >= TIER_A["max_dd"]:
        tier = "A"
    elif sharpe >= TIER_B["sharpe"] and mdd >= TIER_B["max_dd"]:
        tier = "B"
    else:
        tier = "Skip"

    return (tier, score)


# ══════════════════════════════════════════════════════════════════════════════
#  REGIME-AWARE VERDICT  (Phase 4 layer)
# ══════════════════════════════════════════════════════════════════════════════
def compute_today_verdict(asset: str, best_strategy_long: str,
                          ohlc: pd.DataFrame) -> dict:
    """
    Look at today's regime for ``asset`` and return a verdict on whether the
    chosen strategy is allowed to trade RIGHT NOW.

    Returns
    -------
    dict with keys:
        Current Trend    : Bull | Sideways | Bear | "—"
        Current Vol      : Low | Normal | High | "—"
        Allowed Regimes  : human-readable string, e.g. "Bull" or "Bull, Sideways"
        Today's Verdict  : "TRADE" | "STAND DOWN" | "—"
    """
    blank = {
        "Current Trend":   "—",
        "Current Vol":     "—",
        "Allowed Regimes": "—",
        "Today's Verdict": "—",
    }
    if ohlc is None or ohlc.empty:
        return blank

    # Need OHL+Close for vol regime.
    needed = {"High", "Low", "Close"}
    if not needed.issubset(set(ohlc.columns)):
        return blank

    try:
        labelled = label_regimes(ohlc)
        trend_now = labelled["Trend"].dropna()
        vol_now   = labelled["Vol"].dropna()
        if trend_now.empty:
            return blank
        cur_trend = trend_now.iloc[-1]
        cur_vol   = vol_now.iloc[-1] if not vol_now.empty else "—"
    except Exception:
        return blank

    short = LONG_TO_SHORT.get(best_strategy_long)
    if short is None:
        return {
            "Current Trend":   cur_trend,
            "Current Vol":     cur_vol,
            "Allowed Regimes": "—",
            "Today's Verdict": "—",
        }

    allowed = get_allowed_regimes(asset, short)
    allowed_str = ", ".join(sorted(allowed)) if allowed else "(no gate)"

    if not allowed:
        verdict = "TRADE"   # no gate configured → don't block
    else:
        verdict = "TRADE" if cur_trend in allowed else "STAND DOWN"

    return {
        "Current Trend":   cur_trend,
        "Current Vol":     cur_vol,
        "Allowed Regimes": allowed_str,
        "Today's Verdict": verdict,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  MARKET SCREENER — single market pass
# ══════════════════════════════════════════════════════════════════════════════
def screen_market(market_name: str, assets: list, price_data: dict,
                  ohlc_data: dict) -> pd.DataFrame:
    """
    price_data : {asset: Close series}     — drives strategy backtest
    ohlc_data  : {asset: full OHLC frame}  — drives today's regime verdict
    """
    today = dt.datetime.now().strftime("%Y-%m-%d")
    results = []

    print(f"\n  ── Screening {market_name} ({len([a for a in assets if a in price_data])} valid assets) ──")

    for asset in assets:
        if asset not in price_data:
            continue

        full  = price_data[asset]
        train = full[full.index < TRAIN_END]
        test  = full[full.index >= TRAIN_END]

        if len(train) < 100 or len(test) < 50:
            continue

        # ── Best strategy search (train period) ──────────────────────────────
        best_sharpe = -999
        best_strat  = None
        best_params = None

        for strat_name, fn in STRATEGY_FNS.items():
            for params in STRATEGY_GRIDS[strat_name]:
                try:
                    pos = fn(train, **params)
                    m   = compute_metrics(train, pos)
                    if not np.isnan(m["Sharpe"]) and m["Sharpe"] > best_sharpe:
                        best_sharpe = m["Sharpe"]
                        best_strat  = strat_name
                        best_params = params
                except Exception:
                    continue

        if best_strat is None:
            continue

        # ── Out-of-sample (test period) ──────────────────────────────────────
        pos_test = STRATEGY_FNS[best_strat](test, **best_params)
        tm       = compute_metrics(test, pos_test)

        # ── B&H benchmark ────────────────────────────────────────────────────
        log      = np.log(test / test.shift(1)).dropna()
        bah_ret  = (np.exp(log.cumsum()).iloc[-1] - 1) * 100 if len(log) > 0 else np.nan
        bah_cum  = np.exp(log.cumsum()) if len(log) > 0 else pd.Series(dtype=float)
        bah_mdd  = ((bah_cum - bah_cum.cummax()) / bah_cum.cummax()).min() * 100 \
                   if len(bah_cum) > 0 else np.nan
        dd_saved = round(tm["MaxDD"] - float(bah_mdd), 2) \
                   if not np.isnan(bah_mdd) else np.nan

        # ── Phase 4 verdict layer ────────────────────────────────────────────
        verdict = compute_today_verdict(asset, best_strat, ohlc_data.get(asset))

        row = {
            "Market":             market_name,
            "Asset":              asset,
            "Best Strategy":      best_strat,
            "Best Params":        str(best_params),
            "Train Sharpe":       round(best_sharpe, 3),
            "OUT Sharpe":         tm["Sharpe"],
            "OUT Win Rate %":     tm["WinRate"],
            "OUT Strat Ret %":    tm["TotalRet"],
            "OUT B&H Ret %":      round(float(bah_ret), 2) if not np.isnan(bah_ret) else np.nan,
            "OUT Strat Max DD %": tm["MaxDD"],
            "OUT B&H Max DD %":   round(float(bah_mdd), 2) if not np.isnan(bah_mdd) else np.nan,
            "DD Saved %":         dd_saved,
            "Beats B&H":          (tm["TotalRet"] or 0) > (bah_ret or 0),
            # ── Phase 4 columns ──
            "Current Trend":      verdict["Current Trend"],
            "Current Vol":        verdict["Current Vol"],
            "Allowed Regimes":    verdict["Allowed Regimes"],
            "Today's Verdict":    verdict["Today's Verdict"],
            "Run Date":           today,
        }

        tier, score = score_asset(row)
        row["Tier"]  = tier
        row["Score"] = score
        results.append(row)

        # Console output
        tier_icon = {"S": "⭐", "A": "✅", "B": "🔵", "Skip": "⬜"}.get(tier, "")
        beats_icon = "✓" if row["Beats B&H"] else "✗"
        dd_str     = f"{dd_saved:+.1f}%" if not np.isnan(dd_saved) else "N/A"
        v_text     = verdict["Today's Verdict"]
        cur_trend  = verdict["Current Trend"]
        v_icon     = {"TRADE": "🟢", "STAND DOWN": "🔴", "—": "  "}.get(v_text, "  ")
        print(f"    {tier_icon} [{tier}] {asset:14} → {best_strat:18} | "
              f"Sharpe: {tm['Sharpe']:5.2f} | "
              f"Ret: {(tm['TotalRet'] or 0):6.0f}% | "
              f"DD: {(tm['MaxDD'] or 0):5.1f}% | "
              f"DDsaved: {dd_str:>8} | B&H {beats_icon} | Score: {score:.0f}/100 | "
              f"{v_icon} {cur_trend:8} → {v_text}")

    return pd.DataFrame(results)


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL REPORTER
# ══════════════════════════════════════════════════════════════════════════════
TIER_COLORS = {
    "S":    "FFD700",   # Gold
    "A":    "90EE90",   # Light green
    "B":    "87CEEB",   # Sky blue
    "Skip": "D3D3D3",   # Light grey
}

STRAT_COLORS = {
    "MA Crossover":    "AED6F1",
    "RSI":             "FAD7A0",
    "Bollinger Bands": "A9DFBF",
    "MACD":            "D7BDE2",
    "Mean Reversion":  "F1948A",
}

TREND_COLORS = {
    "Bull":     "D5F5E3",  # green
    "Sideways": "FCF3CF",  # yellow
    "Bear":     "FADBD8",  # red
}

VERDICT_COLORS = {
    "TRADE":       "27AE60",
    "STAND DOWN":  "C0392B",
}

def apply_sheet_formatting(ws, df):
    if not EXCEL_FORMAT:
        return
    try:
        THIN   = Side(style="thin", color="CCCCCC")
        BRD    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        HDR_F  = PatternFill("solid", fgColor="1C2833")
        GRN    = PatternFill("solid", fgColor="D5F5E3")
        RED    = PatternFill("solid", fgColor="FADBD8")

        headers = [c.value for c in ws[1]]
        col_idx = {h: i+1 for i, h in enumerate(headers)}

        # Header row
        for cell in ws[1]:
            cell.fill      = HDR_F
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = BRD
        ws.row_dimensions[1].height = 30

        # Data rows
        for row in ws.iter_rows(min_row=2):
            strat_val   = row[col_idx["Best Strategy"]-1].value if "Best Strategy" in col_idx else None
            tier_val    = row[col_idx["Tier"]-1].value          if "Tier"          in col_idx else None
            beats_val   = row[col_idx["Beats B&H"]-1].value     if "Beats B&H"     in col_idx else None
            trend_val   = row[col_idx["Current Trend"]-1].value if "Current Trend" in col_idx else None
            verdict_val = row[col_idx["Today's Verdict"]-1].value if "Today's Verdict" in col_idx else None

            for cell in row:
                cell.border    = BRD
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font      = Font(size=10)
                h = headers[cell.column - 1]

                if h == "Best Strategy" and strat_val:
                    cell.fill = PatternFill("solid", fgColor=STRAT_COLORS.get(strat_val, "FFFFFF"))
                    cell.font = Font(bold=True, size=10)
                elif h == "Tier" and tier_val:
                    cell.fill = PatternFill("solid", fgColor=TIER_COLORS.get(tier_val, "FFFFFF"))
                    cell.font = Font(bold=True, size=10)
                elif h == "Score":
                    try:
                        v = float(cell.value or 0)
                        if v >= 70:   cell.fill = PatternFill("solid", fgColor="27AE60")
                        elif v >= 50: cell.fill = PatternFill("solid", fgColor="F39C12")
                        elif v >= 30: cell.fill = PatternFill("solid", fgColor="AED6F1")
                        else:         cell.fill = PatternFill("solid", fgColor="E8E8E8")
                    except: pass
                elif h == "OUT Sharpe":
                    try:
                        v = float(cell.value or 0)
                        if v >= 0.8:  cell.fill = GRN
                        elif v < 0:   cell.fill = RED
                    except: pass
                elif h == "OUT Strat Ret %":
                    try:
                        v = float(cell.value or 0)
                        if v >= 30:   cell.fill = GRN
                        elif v < -10: cell.fill = RED
                    except: pass
                elif h in ("OUT Strat Max DD %", "OUT B&H Max DD %"):
                    try:
                        v = float(cell.value or 0)
                        if v >= -20:  cell.fill = GRN
                        elif v < -40: cell.fill = RED
                    except: pass
                elif h == "DD Saved %":
                    try:
                        v = float(cell.value or 0)
                        if v >= 10:  cell.fill = GRN
                        elif v < 0:  cell.fill = RED
                    except: pass
                elif h == "Beats B&H":
                    cell.fill = GRN if beats_val else RED
                elif h == "Current Trend" and trend_val in TREND_COLORS:
                    cell.fill = PatternFill("solid", fgColor=TREND_COLORS[trend_val])
                    cell.font = Font(bold=True, size=10)
                elif h == "Today's Verdict" and verdict_val in VERDICT_COLORS:
                    cell.fill = PatternFill("solid", fgColor=VERDICT_COLORS[verdict_val])
                    cell.font = Font(bold=True, color="FFFFFF", size=10)

        # Column widths
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 32)

        # Freeze header
        ws.freeze_panes = "A2"

    except Exception as e:
        print(f"  [WARN] Formatting failed: {e}")


def write_excel_report(all_results: dict, today: str) -> str:
    """Write multi-tab Excel report with per-market tabs + summary."""
    fname = f"Goofy_Phase4_{today}.xlsx"
    path  = os.path.join(OUTPUT_DIR, fname)

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    all_df_list = []
    market_flags = {"US": "🇺🇸", "ASX": "🇦🇺", "JPX": "🇯🇵"}

    # ── Per-market tabs ───────────────────────────────────────────────────────
    for market_name, df in all_results.items():
        if df.empty:
            continue
        df_sorted = df.sort_values("Score", ascending=False).reset_index(drop=True)
        flag = market_flags.get(market_name, "")
        ws = wb.create_sheet(title=f"{flag} {market_name}")

        cols_to_show = [c for c in df_sorted.columns if c != "Best Params"]
        ws.append(cols_to_show)
        for _, row in df_sorted[cols_to_show].iterrows():
            ws.append([row[c] for c in cols_to_show])

        apply_sheet_formatting(ws, df_sorted[cols_to_show])
        all_df_list.append(df_sorted)

    # ── Top Performers tab ────────────────────────────────────────────────────
    if all_df_list:
        combined   = pd.concat(all_df_list, ignore_index=True)
        top_picks  = combined[combined["Tier"].isin(["S", "A"])].sort_values(
                        "Score", ascending=False).reset_index(drop=True)
        decent     = combined[combined["Tier"] == "B"].sort_values(
                        "Score", ascending=False).reset_index(drop=True)

        ws_top = wb.create_sheet(title="⭐ Top Performers", index=0)
        ws_top.append(["GOOFY SCREENER — PHASE 4 RESULTS"])
        ws_top.append([f"Run: {today}  |  Markets: US + ASX + JPX  |  "
                       f"Assets screened: {len(combined)}"])
        ws_top.append([])

        # Tier legend
        ws_top.append(["TIER", "CRITERIA"])
        ws_top.append(["S (⭐ Excellent)", "Sharpe ≥ 0.8, Return ≥ 30%, Max DD ≥ -20%"])
        ws_top.append(["A (✅ Good)",       "Sharpe ≥ 0.4, Return ≥ 10%, Max DD ≥ -35%"])
        ws_top.append(["B (🔵 Decent)",    "Sharpe ≥ 0.1, Max DD ≥ -50%"])
        ws_top.append(["Skip (⬜)",         "Below all thresholds"])
        ws_top.append([])

        # Verdict legend
        ws_top.append(["VERDICT", "MEANING"])
        ws_top.append(["TRADE", "Today's regime is on the strategy's allow-list — green light"])
        ws_top.append(["STAND DOWN", "Today's regime is OUTSIDE the allow-list — sit out today"])
        ws_top.append(["—", "No regime info or no gate configured — strategy runs ungated"])
        ws_top.append([])

        # Summary counts
        ws_top.append(["TIER BREAKDOWN"])
        for tier in ["S", "A", "B", "Skip"]:
            count = len(combined[combined["Tier"] == tier])
            ws_top.append([tier, count])
        ws_top.append([])

        # Verdict counts
        ws_top.append(["VERDICT BREAKDOWN (today)"])
        for v in ["TRADE", "STAND DOWN", "—"]:
            count = len(combined[combined["Today's Verdict"] == v])
            ws_top.append([v, count])
        ws_top.append([])

        ws_top.append(["── S & A TIER (Best Opportunities) ──"])
        if not top_picks.empty:
            cols_show = ["Market", "Asset", "Best Strategy", "Tier", "Score",
                         "OUT Sharpe", "OUT Strat Ret %", "OUT B&H Ret %",
                         "OUT Strat Max DD %", "DD Saved %", "Beats B&H",
                         "Current Trend", "Today's Verdict"]
            ws_top.append(cols_show)
            for _, row in top_picks[cols_show].iterrows():
                ws_top.append([row[c] for c in cols_show])
        else:
            ws_top.append(["No S or A tier assets found this run."])

        ws_top.append([])
        ws_top.append(["── B TIER (Decent — worth watching) ──"])
        if not decent.empty:
            cols_show = ["Market", "Asset", "Best Strategy", "Score",
                         "OUT Sharpe", "OUT Strat Ret %", "OUT Strat Max DD %", "DD Saved %",
                         "Current Trend", "Today's Verdict"]
            ws_top.append(cols_show)
            for _, row in decent[cols_show].iterrows():
                ws_top.append([row[c] for c in cols_show])
        else:
            ws_top.append(["No B tier assets found."])

        # Style the summary sheet header
        if EXCEL_FORMAT:
            try:
                ws_top["A1"].font = Font(bold=True, size=14, color="1C2833")
                ws_top["A2"].font = Font(italic=True, size=10, color="555555")
            except: pass

    # ── Today's Trade List tab (Phase 4 NEW) ──────────────────────────────────
    if all_df_list:
        combined_all = pd.concat(all_df_list, ignore_index=True)
        trade_today  = combined_all[
            (combined_all["Today's Verdict"] == "TRADE") &
            (combined_all["Tier"].isin(["S", "A", "B"]))
        ].sort_values("Score", ascending=False).reset_index(drop=True)

        ws_tt = wb.create_sheet(title="🟢 Today's Trade List", index=1)
        ws_tt.append(["GOOFY SCREENER — TODAY'S TRADE LIST (Phase 4)"])
        ws_tt.append([f"Run: {today}  |  Showing only TRADE-verdict S/A/B tier assets"])
        ws_tt.append([])

        if not trade_today.empty:
            cols = ["Market", "Asset", "Best Strategy", "Tier", "Score",
                    "OUT Sharpe", "OUT Strat Ret %", "OUT Strat Max DD %",
                    "Current Trend", "Current Vol", "Allowed Regimes",
                    "Today's Verdict"]
            ws_tt.append(cols)
            for _, row in trade_today[cols].iterrows():
                ws_tt.append([row[c] for c in cols])

            # Style header
            if EXCEL_FORMAT:
                try:
                    for cell in ws_tt[4]:
                        cell.font = Font(bold=True, color="FFFFFF", size=10)
                        cell.fill = PatternFill("solid", fgColor="1C2833")
                except: pass
        else:
            ws_tt.append(["No assets passed the verdict filter today. "
                          "Either no gates fired, or all gated strategies are STAND DOWN."])

        if EXCEL_FORMAT:
            try:
                ws_tt["A1"].font = Font(bold=True, size=14, color="27AE60")
                ws_tt["A2"].font = Font(italic=True, size=10, color="555555")
            except: pass

    # ── Strategy Distribution tab ─────────────────────────────────────────────
    if all_df_list:
        ws_dist = wb.create_sheet(title="📊 Strategy Distribution")

        ws_dist.append(["Strategy", "US Count", "ASX Count", "JPX Count", "Total"])
        for strat in STRATEGY_FNS.keys():
            row_data = [strat]
            total = 0
            for market in ["US", "ASX", "JPX"]:
                if market in all_results and not all_results[market].empty:
                    c = (all_results[market]["Best Strategy"] == strat).sum()
                else:
                    c = 0
                row_data.append(c)
                total += c
            row_data.append(total)
            ws_dist.append(row_data)

        ws_dist.append([])
        ws_dist.append(["Asset-Strategy Fit (from Phase 1-2 research):"])
        ws_dist.append(["NVDA/TSLA", "→ MACD (momentum)"])
        ws_dist.append(["Sony (6758.T)", "→ Bollinger Bands (mean-reverting)"])
        ws_dist.append(["SPY/ETFs", "→ MA Crossover (efficient markets)"])
        ws_dist.append(["Banks (CBA, MUFG)", "→ RSI / BB"])

        if EXCEL_FORMAT:
            try:
                for cell in ws_dist[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="1C2833")
            except: pass

    # ── Active Gates tab (Phase 4 NEW) ────────────────────────────────────────
    ws_gates = wb.create_sheet(title="🚪 Active Gates")
    ws_gates.append(["ACTIVE ASSET-SPECIFIC GATES"])
    ws_gates.append([f"Loaded from asset_specific_gates.json on run {today}"])
    ws_gates.append([])
    ws_gates.append(["Asset", "Strategy", "Allowed Regimes"])
    if ASSET_SPECIFIC_GATES:
        for asset, strats in ASSET_SPECIFIC_GATES.items():
            for strat, regs in strats.items():
                ws_gates.append([asset, strat, ", ".join(sorted(regs))])
    else:
        ws_gates.append(["—", "—", "(no asset-specific gates loaded; defaults in use)"])

    if EXCEL_FORMAT:
        try:
            ws_gates["A1"].font = Font(bold=True, size=14, color="1C2833")
            ws_gates["A2"].font = Font(italic=True, size=10, color="555555")
            for cell in ws_gates[4]:
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.fill = PatternFill("solid", fgColor="1C2833")
        except: pass

    wb.save(path)
    return path


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="Goofy Phase 4 Multi-Market Screener (regime-aware)")
    parser.add_argument("--market", choices=["US", "ASX", "JPX", "ALL"],
                        default="ALL", help="Which market to screen (default: ALL)")
    args = parser.parse_args()

    today = dt.datetime.now().strftime("%Y-%m-%d")

    # ── Load empirical asset-specific gates if available ──────────────────────
    loaded_gates = load_asset_gates()
    if loaded_gates:
        print(f"  [Phase 4] Loaded {sum(len(v) for v in loaded_gates.values())} "
              f"asset-specific gates across {len(loaded_gates)} assets "
              f"from asset_specific_gates.json")
    else:
        print(f"  [Phase 4] No asset_specific_gates.json found — using "
              f"theory defaults from regime_detector.DEFAULT_STRATEGY_GATES")

    markets_to_run = (["US", "ASX", "JPX"] if args.market == "ALL"
                      else [args.market])

    print(f"""
╔══════════════════════════════════════════════════════════════════════╗
║  GOOFY SCREENER — PHASE 4  |  {today}                    ║
║  Regime-aware verdict layer ON TOP OF Phase 3 strategy selection     ║
║  Markets: {', '.join(markets_to_run):52} ║
║  Train: 2016-2021  |  Test: 2021-{today[:4]}                        ║
╚══════════════════════════════════════════════════════════════════════╝""")

    # ── Step 1: Download all data ─────────────────────────────────────────────
    all_assets = []
    for m in markets_to_run:
        all_assets.extend(UNIVERSE_MAP[m])
    # deduplicate
    seen = set(); unique_assets = []
    for a in all_assets:
        if a not in seen: unique_assets.append(a); seen.add(a)

    print(f"\n[1/3] Downloading data for {len(unique_assets)} assets...\n")
    price_data = {}   # Close series  → for backtest
    ohlc_data  = {}   # full OHLC frame → for regime verdict
    for asset in unique_assets:
        try:
            raw = yf.download(asset, start=TRAIN_START, end=TEST_END,
                              auto_adjust=True, progress=False)
            if not raw.empty and len(raw) >= MIN_ROWS:
                # Flatten yfinance MultiIndex columns if present.
                if isinstance(raw.columns, pd.MultiIndex):
                    raw.columns = raw.columns.get_level_values(0)
                close = raw["Close"].squeeze()
                if isinstance(close, pd.DataFrame):
                    close = close.iloc[:, 0]
                price_data[asset] = close
                # Keep OHLC frame for regime detection (needs High/Low/Close).
                if {"High", "Low", "Close"}.issubset(set(raw.columns)):
                    ohlc_data[asset] = raw[["Open", "High", "Low", "Close"]].copy() \
                        if "Open" in raw.columns else raw[["High", "Low", "Close"]].copy()
                print(f"  ✓ {asset:14} | {len(raw)} rows")
            else:
                print(f"  ✗ {asset:14} | skipped (only {len(raw)} rows)")
        except Exception as e:
            print(f"  ✗ {asset:14} | error: {e}")

    print(f"\n  → {len(price_data)} assets ready for screening "
          f"({len(ohlc_data)} with OHLC for regime detection).\n")

    # ── Step 2: Run screener per market ───────────────────────────────────────
    print(f"[2/3] Running strategy screening + regime verdict...\n")
    all_results = {}
    for m in markets_to_run:
        all_results[m] = screen_market(m, UNIVERSE_MAP[m], price_data, ohlc_data)

    # ── Step 3: Save & summarise ──────────────────────────────────────────────
    print(f"\n[3/3] Saving report...\n")

    all_dfs = [df for df in all_results.values() if not df.empty]
    if not all_dfs:
        print("  No results to save.")
        return

    combined = pd.concat(all_dfs, ignore_index=True)

    # ── Console Summary ────────────────────────────────────────────────────────
    print(f"\n{'═'*70}")
    print(f"  GOOFY SCREENER — PHASE 4 SUMMARY  |  {today}")
    print(f"{'═'*70}")
    print(f"  Assets screened:  {len(combined)}")
    for market in markets_to_run:
        df_m = all_results.get(market, pd.DataFrame())
        if not df_m.empty:
            print(f"    {market:5}: {len(df_m)} assets")

    print(f"\n  ── Tier Breakdown ──")
    for tier in ["S", "A", "B", "Skip"]:
        icon  = {"S":"⭐","A":"✅","B":"🔵","Skip":"⬜"}.get(tier,"")
        count = len(combined[combined["Tier"] == tier])
        bar   = "█" * count
        print(f"    {icon} {tier:5}: {count:3}  {bar}")

    print(f"\n  ── Today's Verdict Breakdown (Phase 4) ──")
    for v in ["TRADE", "STAND DOWN", "—"]:
        icon = {"TRADE":"🟢","STAND DOWN":"🔴","—":"⬜"}.get(v,"")
        count = len(combined[combined["Today's Verdict"] == v])
        bar   = "█" * count
        print(f"    {icon} {v:11}: {count:3}  {bar}")

    print(f"\n  ── 🟢 Today's TRADE List (S/A/B tier only) ──")
    trade_list = combined[
        (combined["Today's Verdict"] == "TRADE") &
        (combined["Tier"].isin(["S","A","B"]))
    ].sort_values("Score", ascending=False)
    if not trade_list.empty:
        for _, r in trade_list.head(15).iterrows():
            print(f"    {r['Market']:5} {r['Asset']:14} | {r['Best Strategy']:18} | "
                  f"Tier {r['Tier']:4} | Score {r['Score']:.0f} | "
                  f"Trend {r['Current Trend']:8} | Sharpe {r['OUT Sharpe']:5.2f}")
    else:
        print("    (no TRADE-verdict S/A/B assets today)")

    print(f"\n  ── 🔴 STAND DOWN — Strategy works historically, but today's regime is wrong ──")
    standdown = combined[
        (combined["Today's Verdict"] == "STAND DOWN") &
        (combined["Tier"].isin(["S","A","B"]))
    ].sort_values("Score", ascending=False)
    if not standdown.empty:
        for _, r in standdown.head(10).iterrows():
            print(f"    {r['Market']:5} {r['Asset']:14} | {r['Best Strategy']:18} | "
                  f"Tier {r['Tier']:4} | Trend now {r['Current Trend']:8} | "
                  f"Allowed {r['Allowed Regimes']}")
    else:
        print("    (none — every S/A/B asset is currently in an allowed regime)")

    print(f"\n  ── ⭐ S-Tier (Excellent) ──")
    s_tier = combined[combined["Tier"] == "S"].sort_values("Score", ascending=False)
    if not s_tier.empty:
        for _, r in s_tier.iterrows():
            v_text = r["Today's Verdict"]
            print(f"    {r['Market']:5} {r['Asset']:14} | {r['Best Strategy']:18} | "
                  f"Sharpe: {r['OUT Sharpe']:5.2f} | Ret: {r['OUT Strat Ret %']:6.1f}% | "
                  f"DD: {r['OUT Strat Max DD %']:5.1f}% | Score: {r['Score']:.0f}/100 | "
                  f"Verdict: {v_text}")
    else:
        print("    (none this run)")

    print(f"\n  ── ✅ A-Tier (Good) ──")
    a_tier = combined[combined["Tier"] == "A"].sort_values("Score", ascending=False)
    if not a_tier.empty:
        for _, r in a_tier.iterrows():
            v_text = r["Today's Verdict"]
            print(f"    {r['Market']:5} {r['Asset']:14} | {r['Best Strategy']:18} | "
                  f"Sharpe: {r['OUT Sharpe']:5.2f} | Ret: {r['OUT Strat Ret %']:6.1f}% | "
                  f"DD: {r['OUT Strat Max DD %']:5.1f}% | Score: {r['Score']:.0f}/100 | "
                  f"Verdict: {v_text}")
    else:
        print("    (none this run)")

    print(f"\n  ── Strategy Distribution ──")
    for strat, count in combined["Best Strategy"].value_counts().items():
        pct = count / len(combined) * 100
        bar = "█" * int(pct / 5)
        print(f"    {strat:20}: {count:3}  ({pct:.0f}%)  {bar}")

    print(f"\n  ── Active Asset-Specific Gates ──")
    if ASSET_SPECIFIC_GATES:
        for asset, strats in ASSET_SPECIFIC_GATES.items():
            for strat, regs in strats.items():
                print(f"    {asset:8} {strat:14} → {sorted(regs)}")
    else:
        print("    (no asset-specific gates loaded — using DEFAULT_STRATEGY_GATES)")

    # ── Save Excel ─────────────────────────────────────────────────────────────
    xlsx_path = write_excel_report(all_results, today)
    print(f"\n  Report saved → {os.path.basename(xlsx_path)}")
    print(f"  Full path:  {xlsx_path}")
    print(f"\n{'═'*70}\n")


if __name__ == "__main__":
    main()
