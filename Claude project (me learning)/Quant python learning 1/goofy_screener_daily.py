"""
Goofy Screener — Daily Automated Script
Runs all 5 strategies across the ASX universe,
finds the best strategy per asset, saves Excel report.

Run manually:  python goofy_screener_daily.py
Scheduled:     runs automatically via Cowork scheduler
"""

import yfinance as yf
import numpy as np
import pandas as pd
import datetime as dt
import os
import warnings
warnings.filterwarnings("ignore")

# ── Try importing Excel formatter (optional) ───────────────────────────────────
try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_FORMAT = True
except ImportError:
    EXCEL_FORMAT = False

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIG
# ══════════════════════════════════════════════════════════════════════════════
UNIVERSE     = "ASX"   # "ASX", "US", or "CUSTOM"
TRAIN_START  = dt.datetime(2016, 1, 1)
TRAIN_END    = dt.datetime(2021, 1, 1)
TEST_END     = dt.datetime.now()
MIN_ROWS     = 500
PERIODS_PER_YEAR = 252

# Output directory — saves inside the workspace folder
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR  = os.path.join(SCRIPT_DIR, "screener_output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Stock universes ────────────────────────────────────────────────────────────
ASX_UNIVERSE = [
    "CBA.AX", "WBC.AX", "ANZ.AX", "NAB.AX",
    "BHP.AX", "RIO.AX", "FMG.AX", "S32.AX",
    "WDS.AX", "STO.AX",
    "CSL.AX", "RMD.AX", "COH.AX",
    "WES.AX", "WOW.AX", "COL.AX",
    "XRO.AX", "WTC.AX",
    "TCL.AX", "APA.AX",
    "IOZ.AX", "STW.AX", "VAS.AX",
]

US_UNIVERSE = [
    "AAPL", "MSFT", "NVDA", "GOOGL", "META", "AMZN", "TSLA",
    "JPM", "BAC", "V", "MA",
    "JNJ", "UNH", "PFE",
    "XOM", "CVX",
    "PG", "KO", "WMT",
    "SPY", "QQQ", "GLD",
]

CUSTOM_UNIVERSE = ["NVDA", "SPY", "CBA.AX", "BHP.AX", "IOZ.AX"]

UNIVERSE_MAP = {"ASX": ASX_UNIVERSE, "US": US_UNIVERSE, "CUSTOM": CUSTOM_UNIVERSE}
ASSETS = UNIVERSE_MAP.get(UNIVERSE, ASX_UNIVERSE)


# ══════════════════════════════════════════════════════════════════════════════
#  STRATEGY FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════
def compute_metrics(price, position):
    df = pd.DataFrame({"price": price})
    df["pos"]      = position.reindex(df.index).fillna(0)
    df["log_ret"]  = np.log(df["price"] / df["price"].shift(1))
    df["strat_ret"] = df["pos"] * df["log_ret"]
    df["cum_strat"] = np.exp(df["strat_ret"].cumsum())

    lr = df["strat_ret"].dropna()
    if len(lr) < 10:
        return {"Sharpe": np.nan, "Total Ret %": np.nan, "Max DD %": np.nan}

    ann   = np.exp(lr.mean() * PERIODS_PER_YEAR) - 1
    vol   = lr.std() * np.sqrt(PERIODS_PER_YEAR)
    sharpe = ann / vol if vol != 0 else np.nan
    cum   = df["cum_strat"].dropna()
    mdd   = ((cum - cum.cummax()) / cum.cummax()).min() if len(cum) > 0 else np.nan
    total = cum.iloc[-1] - 1 if len(cum) > 0 else np.nan

    return {
        "Sharpe":      round(float(sharpe), 3),
        "Total Ret %": round(float(total) * 100, 2),
        "Max DD %":    round(float(mdd) * 100, 2),
    }

def strategy_ma(p, short=20, long=50):
    return (p.rolling(short).mean() > p.rolling(long).mean()).astype(int).shift(1)

def strategy_rsi(p, period=14, oversold=30, overbought=70):
    d = p.diff()
    rsi = 100 - (100 / (1 + d.clip(lower=0).rolling(period).mean() /
                            (-d.clip(upper=0)).rolling(period).mean()))
    sig = np.zeros(len(p)); hold = False
    for i in range(len(rsi)):
        r = rsi.iloc[i]
        if np.isnan(r): sig[i] = 0
        elif not hold and r < oversold:  hold = True;  sig[i] = 1
        elif hold and r > overbought:    hold = False; sig[i] = 0
        else: sig[i] = 1 if hold else 0
    return pd.Series(sig, index=p.index).shift(1)

def strategy_bb(p, window=20, num_std=2.0):
    mid = p.rolling(window).mean()
    low = mid - num_std * p.rolling(window).std()
    sig = np.zeros(len(p)); hold = False
    for i in range(len(p)):
        pi, m, l = p.iloc[i], mid.iloc[i], low.iloc[i]
        if np.isnan(l): sig[i] = 0
        elif not hold and pi <= l: hold = True;  sig[i] = 1
        elif hold and pi >= m:     hold = False; sig[i] = 0
        else: sig[i] = 1 if hold else 0
    return pd.Series(sig, index=p.index).shift(1)

def strategy_macd(p, fast=12, slow=26, signal_p=9):
    macd = p.ewm(span=fast, adjust=False).mean() - p.ewm(span=slow, adjust=False).mean()
    return (macd > macd.ewm(span=signal_p, adjust=False).mean()).astype(int).shift(1)

def strategy_mr(p, window=20, threshold=1.5):
    z = (p - p.rolling(window).mean()) / p.rolling(window).std()
    sig = np.zeros(len(p)); hold = False
    for i in range(len(z)):
        zi = z.iloc[i]
        if np.isnan(zi): sig[i] = 0
        elif not hold and zi < -threshold: hold = True;  sig[i] = 1
        elif hold and zi >= 0:             hold = False; sig[i] = 0
        else: sig[i] = 1 if hold else 0
    return pd.Series(sig, index=p.index).shift(1)

STRATEGY_FNS = {
    "MA Crossover":    strategy_ma,
    "RSI":             strategy_rsi,
    "Bollinger Bands": strategy_bb,
    "MACD":            strategy_macd,
    "Mean Reversion":  strategy_mr,
}

STRATEGY_GRIDS = {
    "MA Crossover":    [{"short": s, "long": l} for s in [10, 20, 50] for l in [50, 100, 200] if s < l],
    "RSI":             [{"period": p, "oversold": os, "overbought": ob} for p in [10, 14, 21] for os, ob in [(25, 65), (30, 70), (35, 75)]],
    "Bollinger Bands": [{"window": w, "num_std": s} for w in [10, 20, 30] for s in [1.5, 2.0, 2.5]],
    "MACD":            [{"fast": f, "slow": s, "signal_p": sig} for f in [8, 12] for s in [21, 26] for sig in [7, 9] if f < s],
    "Mean Reversion":  [{"window": w, "threshold": t} for w in [10, 20, 40] for t in [1.0, 1.5, 2.0]],
}


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN SCREENER
# ══════════════════════════════════════════════════════════════════════════════
def run_screener():
    today = dt.datetime.now().strftime("%Y-%m-%d")
    print(f"\n{'='*60}")
    print(f"  GOOFY SCREENER — Daily Run: {today}")
    print(f"  Universe: {UNIVERSE} ({len(ASSETS)} assets)")
    print(f"{'='*60}\n")

    # Download data
    print("Downloading price data...")
    price_data = {}
    for asset in ASSETS:
        try:
            raw = yf.download(asset, start=TRAIN_START, end=TEST_END,
                              auto_adjust=True, progress=False)
            if not raw.empty and len(raw) >= MIN_ROWS:
                price_data[asset] = raw["Close"].squeeze()
                print(f"  OK {asset}: {len(raw)} rows")
            else:
                print(f"  SKIP {asset}: insufficient data")
        except Exception as e:
            print(f"  ERROR {asset}: {e}")

    valid_assets = list(price_data.keys())
    print(f"\n{len(valid_assets)} assets ready.\n")

    # Run screener
    results = []
    for asset in valid_assets:
        full   = price_data[asset]
        train  = full[full.index < TRAIN_END]
        test   = full[full.index >= TRAIN_END]

        if len(train) < 100 or len(test) < 50:
            continue

        best_sharpe = -999
        best_strat  = None
        best_params = None

        for strat_name, fn in STRATEGY_FNS.items():
            for params in STRATEGY_GRIDS[strat_name]:
                try:
                    pos = fn(train, **params)
                    m   = compute_metrics(train, pos)
                    if not np.isnan(m["Sharpe"]) and m["Sharpe"] > best_sharpe:
                        best_sharpe = m["Sharpe"]
                        best_strat  = strat_name
                        best_params = params
                except:
                    continue

        if best_strat is None:
            continue

        pos  = STRATEGY_FNS[best_strat](test, **best_params)
        tm   = compute_metrics(test, pos)

        log  = np.log(test / test.shift(1)).dropna()
        bah  = (np.exp(log.cumsum()).iloc[-1] - 1) * 100 if len(log) > 0 else np.nan

        # B&H Max Drawdown + DD Saved
        bah_cum  = np.exp(log.cumsum()) if len(log) > 0 else pd.Series(dtype=float)
        bah_mdd  = ((bah_cum - bah_cum.cummax()) / bah_cum.cummax()).min() * 100 if len(bah_cum) > 0 else np.nan
        dd_saved = round(tm["Max DD %"] - float(bah_mdd), 2) if not np.isnan(bah_mdd) else np.nan

        results.append({
            "Asset":              asset,
            "Best Strategy":      best_strat,
            "Train Sharpe":       round(best_sharpe, 3),
            "OUT Sharpe":         tm["Sharpe"],
            "OUT Strat Ret %":    tm["Total Ret %"],
            "OUT B&H Ret %":      round(float(bah), 2),
            "OUT Strat Max DD %": tm["Max DD %"],
            "OUT B&H Max DD %":   round(float(bah_mdd), 2),
            "DD Saved %":         dd_saved,
            "Beats B&H":          tm["Total Ret %"] > bah,
            "Run Date":           today,
        })

        beats    = "✓" if results[-1]["Beats B&H"] else "✗"
        dd_str   = f"{dd_saved:+.1f}%" if not np.isnan(dd_saved) else "N/A"
        print(f"  {asset:12} → {best_strat:18} | "
              f"OUT Sharpe: {tm['Sharpe']:5.2f} | "
              f"Strat DD: {tm['Max DD %']:5.1f}% | B&H DD: {bah_mdd:5.1f}% | Saved: {dd_str} | "
              f"Strat: {tm['Total Ret %']:6.0f}% vs B&H: {bah:6.0f}% {beats}")

    # Save results
    df = pd.DataFrame(results)
    if df.empty:
        print("\nNo results to save.")
        return

    xlsx_path = os.path.join(OUTPUT_DIR, f"Goofy_Screener_{today}.xlsx")
    df.to_excel(xlsx_path, index=False, sheet_name="Screener Results")

    # Apply basic formatting if openpyxl available
    if EXCEL_FORMAT:
        _format_excel(xlsx_path, df, today)

    # Print summary
    print(f"\n{'='*60}")
    print(f"  SUMMARY — {today}")
    print(f"{'='*60}")
    print(f"  Assets screened:   {len(df)}")
    print(f"  Beating B&H:       {df['Beats B&H'].sum()}")
    print(f"\n  Strategy breakdown:")
    for strat, count in df["Best Strategy"].value_counts().items():
        print(f"    {strat:20} {count}")
    print(f"\n  Top 5 by Sharpe:")
    top5 = df.nlargest(5, "OUT Sharpe")[["Asset","Best Strategy","OUT Sharpe","OUT Strat Ret %","OUT Strat Max DD %","DD Saved %"]]
    print(top5.to_string(index=False))
    print(f"\n  Top 5 DD Protectors (DD Saved %):")
    top_dd = df.nlargest(5, "DD Saved %")[["Asset","Best Strategy","OUT Strat Max DD %","OUT B&H Max DD %","DD Saved %"]]
    print(top_dd.to_string(index=False))
    print(f"\n  Report saved: {xlsx_path}")
    print(f"{'='*60}\n")
    return df


def _format_excel(path, df, today):
    try:
        wb = load_workbook(path)
        ws = wb.active
        FILLS = {
            "MA Crossover":    PatternFill("solid", fgColor="AED6F1"),
            "RSI":             PatternFill("solid", fgColor="FAD7A0"),
            "Bollinger Bands": PatternFill("solid", fgColor="A9DFBF"),
            "MACD":            PatternFill("solid", fgColor="D7BDE2"),
            "Mean Reversion":  PatternFill("solid", fgColor="F1948A"),
        }
        HDR = PatternFill("solid", fgColor="2C3E50")
        GRN = PatternFill("solid", fgColor="D5F5E3")
        RED = PatternFill("solid", fgColor="FADBD8")
        THIN = Side(style="thin", color="CCCCCC")
        BRD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        headers    = [c.value for c in ws[1]]
        sc         = headers.index("Best Strategy")      + 1 if "Best Strategy"      in headers else None
        bc         = headers.index("Beats B&H")          + 1 if "Beats B&H"          in headers else None
        oc         = headers.index("OUT Sharpe")         + 1 if "OUT Sharpe"         in headers else None
        mdd_sc     = headers.index("OUT Strat Max DD %") + 1 if "OUT Strat Max DD %" in headers else None
        mdd_bc     = headers.index("OUT B&H Max DD %")   + 1 if "OUT B&H Max DD %"   in headers else None
        dd_save_c  = headers.index("DD Saved %")         + 1 if "DD Saved %"         in headers else None

        for cell in ws[1]:
            cell.fill = HDR
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            cell.border = BRD
        ws.row_dimensions[1].height = 28

        for row in ws.iter_rows(min_row=2):
            sv = row[sc-1].value if sc else None
            bv = row[bc-1].value if bc else None
            for cell in row:
                cell.border = BRD
                cell.alignment = Alignment(horizontal="center")
                if sc and cell.column == sc:
                    cell.fill = FILLS.get(sv, PatternFill())
                    cell.font = Font(bold=True)
                if bc and cell.column == bc:
                    cell.fill = GRN if bv else RED
                if oc and cell.column == oc:
                    try:
                        v = float(cell.value)
                        if v >= 0.5: cell.fill = GRN
                        elif v < 0:  cell.fill = RED
                    except: pass
                if mdd_sc and cell.column == mdd_sc:
                    try:
                        v = float(cell.value)
                        if v >= -15:          cell.fill = GRN
                        elif -30 <= v < -15:  pass  # no fill = yellow-ish
                        elif v < -30:         cell.fill = RED
                    except: pass
                if mdd_bc and cell.column == mdd_bc:
                    try:
                        v = float(cell.value)
                        if v >= -15:  cell.fill = GRN
                        elif v < -30: cell.fill = RED
                    except: pass
                if dd_save_c and cell.column == dd_save_c:
                    try:
                        v = float(cell.value)
                        if v >= 10:  cell.fill = GRN   # saved ≥10% drawdown
                        elif v < 0:  cell.fill = RED   # strategy was worse than holding
                    except: pass

        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 30)
        ws.freeze_panes = "A2"
        wb.save(path)
    except Exception as e:
        print(f"  (Excel formatting skipped: {e})")


if __name__ == "__main__":
    run_screener()
