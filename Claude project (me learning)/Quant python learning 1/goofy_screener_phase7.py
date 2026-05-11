"""
╔══════════════════════════════════════════════════════════════════════╗
║           GOOFY SCREENER — PHASE 7                                   ║
║           ML Signal Layer: XGBoost Directional Classifier            ║
║                                                                      ║
║   Inherits Phase 6 (correlation-aware portfolio) — adds a trained   ║
║   XGBoost model per asset that predicts whether the next 20 days    ║
║   will be directionally favourable.                                  ║
║                                                                      ║
║   Three new columns per asset:                                       ║
║       ML Score    — model probability (0–1) for current conditions   ║
║       ML Gate     — PASS (≥0.55) or HOLD (<0.55)                    ║
║       P7 Verdict  — TRADE / ML HOLD / STAND DOWN                    ║
║                                                                      ║
║   One new Excel tab:                                                 ║
║       🤖 ML Signals — scores, AUC, feature importance per asset     ║
║                                                                      ║
║   Run:  python3 goofy_screener_phase7.py                            ║
║   Args: --market US | ASX | JPX | ALL (default: ALL)               ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import yfinance as yf
import numpy as np
import pandas as pd
import datetime as dt
import os, sys, warnings, argparse, ast
warnings.filterwarnings("ignore")

# ── Phase 5 screening layer ───────────────────────────────────────────────────
from goofy_screener_phase5 import (
    UNIVERSE_MAP, STRATEGY_FNS, STRATEGY_GRIDS,
    TRAIN_START, TRAIN_END, TEST_END, MIN_ROWS,
    TARGET_VOL, KELLY_FRACTION, OUTPUT_DIR,
    compute_metrics, score_asset, compute_today_verdict,
    apply_sheet_formatting, _style_header,
    LONG_TO_SHORT,
)
from position_sizer import compute_trade_stats, recommend_size
from regime_detector import load_asset_gates, ASSET_SPECIFIC_GATES

# ── Phase 6 portfolio layer ───────────────────────────────────────────────────
from portfolio_builder import (
    compute_correlation_matrix, find_clusters,
    adjust_for_correlation, portfolio_metrics,
    cluster_label, CORR_THRESHOLD,
)

# ── Phase 7 ML layer ──────────────────────────────────────────────────────────
from ml_signal import (
    engineer_features, build_ml_model, get_current_score,
    ml_gate, combined_verdict, XGB_AVAILABLE,
    ML_PASS_THRESH, LOOKFORWARD,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    EXCEL_FORMAT = True
except ImportError:
    EXCEL_FORMAT = False

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

VERDICT_COLORS_P7 = {
    "TRADE":      "27AE60",
    "ML HOLD":    "F39C12",
    "STAND DOWN": "E74C3C",
}


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 7: PER-ASSET ML TRAINING (runs inside the screening loop)
# ══════════════════════════════════════════════════════════════════════════════

def run_ml_layer(all_results: dict, price_data: dict,
                 ohlc_data: dict) -> dict:
    """
    For every asset in all_results, train an XGBoost model and attach:
        ML Score, ML Gate, Val AUC, P7 Verdict
    Returns updated all_results dict.
    """
    if not XGB_AVAILABLE:
        print("  [Phase 7] XGBoost not available — skipping ML layer.")
        return all_results

    train_end_str = TRAIN_END.strftime("%Y-%m-%d")
    total = sum(len(df) for df in all_results.values() if not df.empty)
    done  = 0

    for market, df in all_results.items():
        if df.empty:
            continue
        ml_scores  = []
        ml_gates   = []
        val_aucs   = []
        p7_verdicts= []

        for _, row in df.iterrows():
            asset   = row["Asset"]
            verdict = row.get("Today's Verdict", "—")

            price = price_data.get(asset)
            ohlc  = ohlc_data.get(asset)

            if price is None or len(price) < 300:
                ml_scores.append(None); ml_gates.append("PASS")
                val_aucs.append(None); p7_verdicts.append(verdict)
                done += 1; continue

            model, feat_names, train_auc, val_auc = build_ml_model(
                price, ohlc, train_end=train_end_str
            )
            score  = get_current_score(model, feat_names, price, ohlc)
            gate   = ml_gate(score)
            p7_v   = combined_verdict(verdict, gate)

            ml_scores.append(round(score * 100, 1) if score is not None else None)
            ml_gates.append(gate)
            val_aucs.append(round(val_auc, 3) if val_auc is not None else None)
            p7_verdicts.append(p7_v)

            done += 1
            auc_str   = f"AUC={val_auc:.2f}" if val_auc else "AUC=—"
            score_str = f"{score*100:.0f}%" if score else "—"
            icon = "🟢" if p7_v == "TRADE" else "🟡" if p7_v == "ML HOLD" else "🔴"
            print(f"    {icon} {asset:14} | ML={score_str:5} Gate={gate:4} "
                  f"{auc_str}  →  {p7_v}")

        df = df.copy()
        df["ML Score"]  = ml_scores
        df["ML Gate"]   = ml_gates
        df["Val AUC"]   = val_aucs
        df["P7 Verdict"]= p7_verdicts
        all_results[market] = df

    return all_results


# ══════════════════════════════════════════════════════════════════════════════
#  SCREENING (same as Phase 5 — needed for position series access)
# ══════════════════════════════════════════════════════════════════════════════

def screen_market_p7(market_name: str, assets: list,
                     price_data: dict, ohlc_data: dict) -> pd.DataFrame:
    """Phase 5 screening logic — identical output, just called by Phase 7 main."""
    today   = dt.datetime.now().strftime("%Y-%m-%d")
    results = []
    print(f"\n  ── Screening {market_name} "
          f"({len([a for a in assets if a in price_data])} valid assets) ──")

    for asset in assets:
        if asset not in price_data:
            continue
        full  = price_data[asset]
        train = full[full.index < TRAIN_END]
        test  = full[full.index >= TRAIN_END]
        if len(train) < 100 or len(test) < 50:
            continue

        best_sharpe = -999; best_strat = None; best_params = None
        for strat_name, fn in STRATEGY_FNS.items():
            for params in STRATEGY_GRIDS[strat_name]:
                try:
                    pos = fn(train, **params)
                    m   = compute_metrics(train, pos)
                    if not np.isnan(m["Sharpe"]) and m["Sharpe"] > best_sharpe:
                        best_sharpe = m["Sharpe"]
                        best_strat  = strat_name
                        best_params = params
                except Exception:
                    continue
        if best_strat is None:
            continue

        pos_test = STRATEGY_FNS[best_strat](test, **best_params)
        tm       = compute_metrics(test, pos_test)

        log     = np.log(test / test.shift(1)).dropna()
        bah_ret = (np.exp(log.cumsum()).iloc[-1] - 1) * 100 if len(log) > 0 else np.nan
        bah_cum = np.exp(log.cumsum()) if len(log) > 0 else pd.Series(dtype=float)
        bah_mdd = ((bah_cum - bah_cum.cummax()) / bah_cum.cummax()).min() * 100 \
                  if len(bah_cum) > 0 else np.nan
        dd_saved = round(tm["MaxDD"] - float(bah_mdd), 2) if not np.isnan(bah_mdd) else np.nan

        verdict      = compute_today_verdict(asset, best_strat, ohlc_data.get(asset))
        trade_stats  = compute_trade_stats(test, pos_test)
        sizing       = recommend_size(trade_stats, test.pct_change().dropna(),
                                      target_vol=TARGET_VOL,
                                      kelly_fraction=KELLY_FRACTION)

        row = {
            "Market": market_name, "Asset": asset,
            "Best Strategy": best_strat, "Best Params": str(best_params),
            "Train Sharpe": round(best_sharpe, 3),
            "OUT Sharpe": tm["Sharpe"], "OUT Win Rate %": tm["WinRate"],
            "OUT Strat Ret %": tm["TotalRet"],
            "OUT B&H Ret %": round(float(bah_ret), 2) if not np.isnan(bah_ret) else np.nan,
            "OUT Strat Max DD %": tm["MaxDD"],
            "OUT B&H Max DD %":  round(float(bah_mdd), 2) if not np.isnan(bah_mdd) else np.nan,
            "DD Saved %": dd_saved, "Beats B&H": (tm["TotalRet"] or 0) > (bah_ret or 0),
            "Current Trend": verdict["Current Trend"],
            "Current Vol":   verdict["Current Vol"],
            "Allowed Regimes": verdict["Allowed Regimes"],
            "Today's Verdict": verdict["Today's Verdict"],
            "N Trades": trade_stats["n_trades"] if trade_stats else None,
            "Trade Win Rate %": round(trade_stats["win_rate"]*100, 1) if trade_stats else None,
            "Avg Win %":  round(trade_stats["avg_win"]*100, 2)  if trade_stats else None,
            "Avg Loss %": round(trade_stats["avg_loss"]*100, 2) if trade_stats else None,
            "Kelly %": sizing["Kelly %"], "Vol Scalar": sizing["Vol Scalar"],
            "Recommended Size %": sizing["Recommended Size %"],
            "Run Date": today,
        }
        tier, score = score_asset(row)
        row["Tier"] = tier; row["Score"] = score
        results.append(row)

        tier_icon = {"S": "⭐", "A": "✅", "B": "🔵", "Skip": "⬜"}.get(tier, "")
        v_verdict = verdict["Today's Verdict"]
        print(f"    {tier_icon} [{tier}] {asset:14} → {best_strat:18} | "
              f"Sharpe: {tm['Sharpe']:5.2f} | Score: {score:.0f}/100 | {v_verdict}")

    return pd.DataFrame(results)


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL REPORTER (Phase 7)
# ══════════════════════════════════════════════════════════════════════════════

def write_excel_phase7(all_results: dict, today: str, p6: dict) -> str:
    from goofy_screener_phase6 import write_excel_phase6
    # Re-use Phase 6 Excel as base, then reopen and add ML tab
    from openpyxl import load_workbook

    fname_p6 = f"Goofy_Phase6_{today}_tmp.xlsx"
    path_tmp = os.path.join(OUTPUT_DIR, fname_p6)

    # Write Phase 6 base (strip ML cols not in phase6 writer)
    all_p6 = {}
    for m, df in all_results.items():
        if df.empty:
            all_p6[m] = df; continue
        all_p6[m] = df.drop(columns=[c for c in
                    ["ML Score", "ML Gate", "Val AUC", "P7 Verdict"]
                    if c in df.columns], errors="ignore")
    write_excel_phase6(all_p6, today, p6)

    # Now write the real Phase 7 file from scratch (with ML cols)
    fname = f"Goofy_Phase7_{today}.xlsx"
    path  = os.path.join(OUTPUT_DIR, fname)

    wb    = Workbook()
    wb.remove(wb.active)
    market_flags = {"US": "🇺🇸", "ASX": "🇦🇺", "JPX": "🇯🇵"}
    HIDE_COLS    = {"Best Params"}

    all_df_list = []
    for market_name, df in all_results.items():
        if df.empty: continue
        df_sorted = df.sort_values("Score", ascending=False).reset_index(drop=True)
        flag = market_flags.get(market_name, "")
        ws   = wb.create_sheet(title=f"{flag} {market_name}")
        cols = [c for c in df_sorted.columns if c not in HIDE_COLS]
        ws.append(cols)
        for _, row in df_sorted[cols].iterrows():
            ws.append([row[c] for c in cols])
        apply_sheet_formatting(ws, df_sorted[cols])
        # Extra P7 formatting
        if EXCEL_FORMAT:
            try:
                hdr = cols
                for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in r:
                        h = hdr[cell.column-1] if cell.column-1 < len(hdr) else ""
                        if h == "P7 Verdict" and cell.value in VERDICT_COLORS_P7:
                            cell.fill = PatternFill("solid",
                                fgColor=VERDICT_COLORS_P7[cell.value])
                            cell.font = Font(bold=True, color="FFFFFF", size=10)
                        elif h == "ML Score" and cell.value is not None:
                            try:
                                v = float(cell.value)
                                if v >= 65: cell.fill = PatternFill("solid", fgColor="27AE60")
                                elif v >= 55: cell.fill = PatternFill("solid", fgColor="F39C12")
                                else: cell.fill = PatternFill("solid", fgColor="E74C3C")
                            except: pass
            except: pass
        all_df_list.append(df_sorted)

    if not all_df_list:
        wb.save(path); return path

    combined = pd.concat(all_df_list, ignore_index=True)
    adj_sizes    = p6.get("adj_sizes", {})
    cluster_map  = p6.get("cluster_map", {})
    cluster_sizes= p6.get("cluster_sizes", {})
    corr_risk    = p6.get("corr_risk", {})

    for market, df in all_results.items():
        if not df.empty:
            df["Cluster"]    = df["Asset"].map(lambda a: cluster_label(a, cluster_map, cluster_sizes))
            df["Corr Risk"]  = df["Asset"].map(lambda a: corr_risk.get(a, "—"))
            df["Adj Size %"] = df["Asset"].map(lambda a: adj_sizes.get(a))
            all_results[market] = df

    combined = pd.concat([df for df in all_results.values() if not df.empty],
                         ignore_index=True)

    # ── Today's Trade List (P7) ───────────────────────────────────────────────
    trade_p7 = combined[
        (combined["P7 Verdict"] == "TRADE") &
        (combined["Tier"].isin(["S", "A", "B"]))
    ].sort_values("Score", ascending=False).reset_index(drop=True)

    ws_tt = wb.create_sheet(title="🟢 Today's Trade List", index=0)
    ws_tt.append(["GOOFY SCREENER — PHASE 7 TRADE LIST"])
    ws_tt.append([f"Run: {today}  |  P7 Verdict = TRADE (regime gate AND ML gate passed)"])
    ws_tt.append([])
    if not trade_p7.empty:
        cols = ["Market", "Asset", "Best Strategy", "Tier", "Score",
                "OUT Sharpe", "OUT Strat Ret %", "OUT Strat Max DD %",
                "Current Trend", "Today's Verdict",
                "ML Score", "ML Gate", "P7 Verdict",
                "Kelly %", "Recommended Size %", "Adj Size %"]
        cols = [c for c in cols if c in trade_p7.columns]
        ws_tt.append(cols)
        for _, row in trade_p7[cols].iterrows():
            ws_tt.append([row[c] for c in cols])
        _style_header(ws_tt, row_num=4)
        ws_tt.auto_filter.ref = f"A4:{get_column_letter(len(cols))}{ws_tt.max_row}"
        ws_tt.freeze_panes = "A5"

    # ── ML Signals tab ────────────────────────────────────────────────────────
    ws_ml = wb.create_sheet(title="🤖 ML Signals")
    ws_ml.append(["PHASE 7 — ML SIGNAL LAYER"])
    ws_ml.append([f"Run: {today}  |  XGBoost — 20-day directional classifier  |  "
                  f"PASS threshold: {ML_PASS_THRESH*100:.0f}%  |  "
                  f"Features: momentum, RSI, MACD, BB, MA trend, vol, drawdown"])
    ws_ml.append([])
    ws_ml.append(["HOW TO READ"])
    ws_ml.append(["ML Score", f"Probability (0–100%) that conditions favour upward price move "
                               f"over next {LOOKFORWARD} trading days"])
    ws_ml.append(["ML Gate",  f"PASS if ML Score ≥ {ML_PASS_THRESH*100:.0f}%, HOLD otherwise"])
    ws_ml.append(["Val AUC",  "Area Under Curve on test period (2021–present). "
                               "0.50 = random, 0.60+ = some signal, 0.70+ = strong. "
                               "Below 0.50 = model predicts wrong direction."])
    ws_ml.append(["P7 Verdict", "TRADE = regime gate + ML gate both pass. "
                                  "ML HOLD = regime says trade but ML says wait. "
                                  "STAND DOWN = regime gate failed (ML irrelevant)."])
    ws_ml.append([])

    ml_sorted = combined.sort_values("Score", ascending=False).reset_index(drop=True)
    cols_ml = ["Market", "Asset", "Tier", "Score", "Best Strategy",
               "Today's Verdict", "ML Score", "ML Gate", "Val AUC", "P7 Verdict",
               "Kelly %", "Recommended Size %"]
    cols_ml = [c for c in cols_ml if c in ml_sorted.columns]
    ml_header_row = ws_ml.max_row + 1
    ws_ml.append(cols_ml)
    for _, row in ml_sorted[cols_ml].iterrows():
        ws_ml.append([row[c] for c in cols_ml])
    _style_header(ws_ml, row_num=ml_header_row)
    ws_ml.auto_filter.ref = (f"A{ml_header_row}:"
                              f"{get_column_letter(len(cols_ml))}{ws_ml.max_row}")
    ws_ml.freeze_panes = f"A{ml_header_row + 1}"

    if EXCEL_FORMAT:
        try:
            hdr = cols_ml
            for r in ws_ml.iter_rows(min_row=ml_header_row+1, max_row=ws_ml.max_row):
                for cell in r:
                    h = hdr[cell.column-1] if cell.column-1 < len(hdr) else ""
                    if h == "P7 Verdict" and cell.value in VERDICT_COLORS_P7:
                        cell.fill = PatternFill("solid",
                            fgColor=VERDICT_COLORS_P7[cell.value])
                        cell.font = Font(bold=True, color="FFFFFF", size=10)
                    elif h == "ML Score" and cell.value is not None:
                        try:
                            v = float(cell.value)
                            if v >= 65: cell.fill = PatternFill("solid", fgColor="27AE60")
                            elif v >= 55: cell.fill = PatternFill("solid", fgColor="F39C12")
                            else: cell.fill = PatternFill("solid", fgColor="E74C3C")
                        except: pass
            ws_ml["A1"].font = Font(bold=True, size=14, color="1C2833")
            ws_ml["A2"].font = Font(italic=True, size=10, color="555555")
            for col in ws_ml.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                ws_ml.column_dimensions[get_column_letter(col[0].column)].width = min(max_len+4, 36)
        except: pass

    wb.save(path)

    # Clean up temp file
    try:
        p6_file = os.path.join(OUTPUT_DIR, f"Goofy_Phase6_{today}_tmp.xlsx")
        if os.path.exists(p6_file): os.remove(p6_file)
    except: pass

    return path


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Goofy Phase 7 — ML signal layer (XGBoost)")
    parser.add_argument("--market", choices=["US", "ASX", "JPX", "ALL"],
                        default="ALL")
    args  = parser.parse_args()
    today = dt.datetime.now().strftime("%Y-%m-%d")

    loaded_gates = load_asset_gates()
    markets_to_run = ["US", "ASX", "JPX"] if args.market == "ALL" else [args.market]

    print(f"""
╔══════════════════════════════════════════════════════════════════════╗
║  GOOFY SCREENER — PHASE 7  |  {today}                    ║
║  Strategy + Regime + Kelly + Correlation + XGBoost ML Signal        ║
║  Markets: {', '.join(markets_to_run):52} ║
║  ML: 20-day directional classifier  |  PASS threshold: {ML_PASS_THRESH*100:.0f}%       ║
╚══════════════════════════════════════════════════════════════════════╝""")

    # ── Step 1: Download ──────────────────────────────────────────────────────
    all_assets = []
    for m in markets_to_run: all_assets.extend(UNIVERSE_MAP[m])
    seen = set(); unique_assets = []
    for a in all_assets:
        if a not in seen: unique_assets.append(a); seen.add(a)

    print(f"\n[1/5] Downloading {len(unique_assets)} assets...\n")
    price_data = {}; ohlc_data = {}
    for asset in unique_assets:
        try:
            raw = yf.download(asset, start=TRAIN_START, end=TEST_END,
                              auto_adjust=True, progress=False)
            if not raw.empty and len(raw) >= MIN_ROWS:
                if isinstance(raw.columns, pd.MultiIndex):
                    raw.columns = raw.columns.get_level_values(0)
                close = raw["Close"].squeeze()
                if isinstance(close, pd.DataFrame): close = close.iloc[:, 0]
                price_data[asset] = close
                if {"High", "Low", "Close"}.issubset(set(raw.columns)):
                    ohlc_data[asset] = raw[["High", "Low", "Close"]].copy()
                print(f"  ✓ {asset:14} | {len(raw)} rows")
            else:
                print(f"  ✗ {asset:14} | skipped")
        except Exception as e:
            print(f"  ✗ {asset:14} | error: {e}")

    # ── Step 2: Phase 5 screening ─────────────────────────────────────────────
    print(f"\n[2/5] Phase 5 screening (strategy + regime + sizing)...\n")
    all_results = {}
    for m in markets_to_run:
        all_results[m] = screen_market_p7(m, UNIVERSE_MAP[m], price_data, ohlc_data)

    all_dfs = [df for df in all_results.values() if not df.empty]
    if not all_dfs: print("  No results."); return
    combined = pd.concat(all_dfs, ignore_index=True)

    # ── Step 3: Phase 6 portfolio construction ───────────────────────────────
    print(f"\n[3/5] Phase 6 — correlation matrix + cluster sizing...\n")
    from goofy_screener_phase6 import run_portfolio_construction, attach_phase6_columns
    p6 = run_portfolio_construction(combined, price_data)
    all_results = attach_phase6_columns(all_results, p6)
    combined = pd.concat([df for df in all_results.values() if not df.empty],
                         ignore_index=True)

    # ── Step 4: Phase 7 ML layer ──────────────────────────────────────────────
    print(f"\n[4/5] Phase 7 — training XGBoost per asset...\n")
    all_results = run_ml_layer(all_results, price_data, ohlc_data)
    combined = pd.concat([df for df in all_results.values() if not df.empty],
                         ignore_index=True)

    # ── Step 5: Save ──────────────────────────────────────────────────────────
    print(f"\n[5/5] Saving Phase 7 report...\n")

    print(f"\n{'═'*72}")
    print(f"  GOOFY SCREENER — PHASE 7 SUMMARY  |  {today}")
    print(f"{'═'*72}")

    for tier in ["S", "A", "B", "Skip"]:
        print(f"  {['⭐','✅','🔵','⬜'][['S','A','B','Skip'].index(tier)]} {tier}: "
              f"{len(combined[combined['Tier']==tier])}")

    print(f"\n  ── P7 Verdict breakdown ──")
    for v in ["TRADE", "ML HOLD", "STAND DOWN"]:
        if "P7 Verdict" in combined.columns:
            n = len(combined[combined["P7 Verdict"] == v])
            icon = {"TRADE": "🟢", "ML HOLD": "🟡", "STAND DOWN": "🔴"}.get(v, "")
            print(f"    {icon} {v:12}: {n}")

    if "P7 Verdict" in combined.columns:
        trade_p7 = combined[
            (combined["P7 Verdict"] == "TRADE") &
            (combined["Tier"].isin(["S", "A", "B"]))
        ].sort_values("Score", ascending=False)

        if not trade_p7.empty:
            print(f"\n  ── TRADE signals (both gates passed) ──")
            print(f"  {'Asset':14} {'Strategy':18} {'Tier':4} {'Score':6} "
                  f"{'ML%':6} {'AUC':6} {'AdjSize%':9}")
            print(f"  {'─'*70}")
            for _, r in trade_p7.head(20).iterrows():
                ml  = f"{r['ML Score']:.0f}%" if r.get('ML Score') is not None else "—"
                auc = f"{r['Val AUC']:.2f}" if r.get('Val AUC') is not None else "—"
                adj = f"{p6.get('adj_sizes',{}).get(r['Asset'],0):.0f}%"
                print(f"  {r['Asset']:14} {r['Best Strategy']:18} {r['Tier']:4} "
                      f"{r['Score']:5.0f} {ml:>6} {auc:>6} {adj:>9}")

    xlsx_path = write_excel_phase7(all_results, today, p6)
    print(f"\n  Report saved → {os.path.basename(xlsx_path)}")
    print(f"  Full path:   {xlsx_path}")
    print(f"\n{'═'*72}\n")


if __name__ == "__main__":
    main()
