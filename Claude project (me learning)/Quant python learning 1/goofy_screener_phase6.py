"""
╔══════════════════════════════════════════════════════════════════════╗
║           GOOFY SCREENER — PHASE 6                                   ║
║           Portfolio Construction: Correlation + Clustering            ║
║                                                                      ║
║   Inherits Phase 5 (Kelly + vol-scaled sizing) — adds a portfolio   ║
║   layer that asks: "how do these positions interact with each other?"║
║                                                                      ║
║   Three new columns per asset:                                       ║
║       Cluster      — which correlated group this asset belongs to    ║
║       Corr Risk    — Low / Medium / High (cluster size)              ║
║       Adj Size %   — Phase 5 size scaled by 1/√(cluster_size)       ║
║                                                                      ║
║   Two new Excel tabs:                                                ║
║       🔗 Correlation Matrix — pairwise return correlations (heatmap) ║
║       📊 Portfolio View     — final weights + portfolio metrics      ║
║                                                                      ║
║   Run:  python3 goofy_screener_phase6.py                            ║
║   Args: --market US | ASX | JPX | ALL (default: ALL)               ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import yfinance as yf
import numpy as np
import pandas as pd
import datetime as dt
import os, sys, warnings, argparse
warnings.filterwarnings("ignore")

# ── Phase 5 screening layer (all strategy + regime + sizing logic) ────────────
from goofy_screener_phase5 import (
    screen_market,
    UNIVERSE_MAP,
    STRATEGY_FNS,
    TRAIN_START, TRAIN_END, TEST_END, MIN_ROWS,
    TARGET_VOL, KELLY_FRACTION, OUTPUT_DIR,
    apply_sheet_formatting, _style_header,
    LONG_TO_SHORT,
)
from regime_detector import load_asset_gates, ASSET_SPECIFIC_GATES

# ── Phase 6 portfolio construction layer ──────────────────────────────────────
from portfolio_builder import (
    compute_correlation_matrix,
    find_clusters,
    adjust_for_correlation,
    portfolio_metrics,
    cluster_label,
    CORR_THRESHOLD,
)

# ── Optional: rich Excel formatting ───────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    EXCEL_FORMAT = True
except ImportError:
    EXCEL_FORMAT = False

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

CORR_RISK_COLORS = {
    "Low":    "27AE60",   # green
    "Medium": "F39C12",   # orange
    "High":   "E74C3C",   # red
}


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 6: PORTFOLIO CONSTRUCTION
# ══════════════════════════════════════════════════════════════════════════════

def run_portfolio_construction(combined_df: pd.DataFrame,
                               price_data: dict) -> dict:
    """
    Takes Phase 5 combined results + raw price series.
    Returns a dict with all Phase 6 outputs.
    """
    trade_assets = combined_df[
        (combined_df["Today's Verdict"] == "TRADE") &
        (combined_df["Tier"] != "Skip")
    ].copy()

    if trade_assets.empty or len(trade_assets) < 2:
        return {"cluster_map": {}, "adj_sizes": {}, "cluster_sizes": {},
                "corr_risk": {}, "corr_matrix": pd.DataFrame(),
                "port_metrics": {}, "returns_dict": {}}

    # Build returns from test period (2021-present) only
    test_start = pd.Timestamp("2021-01-01")
    returns_dict = {}
    for asset in trade_assets["Asset"]:
        if asset in price_data:
            prices = price_data[asset]
            test_prices = prices[prices.index >= test_start]
            if len(test_prices) >= 60:
                returns_dict[asset] = test_prices.pct_change().dropna()

    if len(returns_dict) < 2:
        return {"cluster_map": {}, "adj_sizes": {}, "cluster_sizes": {},
                "corr_risk": {}, "corr_matrix": pd.DataFrame(),
                "port_metrics": {}, "returns_dict": returns_dict}

    corr_matrix = compute_correlation_matrix(returns_dict)
    cluster_map = find_clusters(corr_matrix, CORR_THRESHOLD)

    raw_sizes = {}
    for _, row in trade_assets.iterrows():
        raw_sizes[row["Asset"]] = row.get("Recommended Size %") or 0.0

    assets_list = list(trade_assets["Asset"])
    adj_sizes, cluster_sizes, corr_risk = adjust_for_correlation(
        assets_list, raw_sizes, cluster_map
    )

    weights    = {a: v / 100.0 for a, v in adj_sizes.items() if v > 0}
    port_m     = portfolio_metrics(returns_dict, weights)

    return {
        "cluster_map":   cluster_map,
        "adj_sizes":     adj_sizes,
        "cluster_sizes": cluster_sizes,
        "corr_risk":     corr_risk,
        "corr_matrix":   corr_matrix,
        "port_metrics":  port_m,
        "returns_dict":  returns_dict,
    }


def attach_phase6_columns(all_results: dict, p6: dict) -> dict:
    """Add Cluster, Corr Risk, Adj Size % columns to each market's DataFrame."""
    cluster_map   = p6.get("cluster_map", {})
    cluster_sizes = p6.get("cluster_sizes", {})
    corr_risk     = p6.get("corr_risk", {})
    adj_sizes     = p6.get("adj_sizes", {})

    for market, df in all_results.items():
        if df.empty:
            continue
        df = df.copy()
        df["Cluster"]    = df["Asset"].map(
            lambda a: cluster_label(a, cluster_map, cluster_sizes))
        df["Corr Risk"]  = df["Asset"].map(lambda a: corr_risk.get(a, "—"))
        df["Adj Size %"] = df["Asset"].map(lambda a: adj_sizes.get(a))
        all_results[market] = df

    return all_results


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL REPORTER (Phase 6)
# ══════════════════════════════════════════════════════════════════════════════

def write_excel_phase6(all_results: dict, today: str, p6: dict) -> str:
    fname = f"Goofy_Phase6_{today}.xlsx"
    path  = os.path.join(OUTPUT_DIR, fname)

    wb = Workbook()
    wb.remove(wb.active)

    all_df_list  = []
    market_flags = {"US": "🇺🇸", "ASX": "🇦🇺", "JPX": "🇯🇵"}
    HIDE_COLS    = {"Best Params"}

    # ── Per-market tabs (Phase 5 formatting + Phase 6 columns) ───────────────
    for market_name, df in all_results.items():
        if df.empty:
            continue
        df_sorted = df.sort_values("Score", ascending=False).reset_index(drop=True)
        flag      = market_flags.get(market_name, "")
        ws        = wb.create_sheet(title=f"{flag} {market_name}")

        cols_to_show = [c for c in df_sorted.columns if c not in HIDE_COLS]
        ws.append(cols_to_show)
        for _, row in df_sorted[cols_to_show].iterrows():
            ws.append([row[c] for c in cols_to_show])

        apply_sheet_formatting(ws, df_sorted[cols_to_show])

        # Extra formatting for Phase 6 columns
        if EXCEL_FORMAT:
            try:
                headers = cols_to_show
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        h = headers[cell.column - 1] if cell.column - 1 < len(headers) else ""
                        if h == "Corr Risk" and cell.value in CORR_RISK_COLORS:
                            cell.fill = PatternFill("solid",
                                fgColor=CORR_RISK_COLORS[cell.value])
                            cell.font = Font(bold=True, color="FFFFFF", size=10)
                        elif h == "Adj Size %" and cell.value is not None:
                            try:
                                v = float(cell.value)
                                if v >= 40:
                                    cell.fill = PatternFill("solid", fgColor="27AE60")
                                elif v >= 15:
                                    cell.fill = PatternFill("solid", fgColor="F39C12")
                                elif v > 0:
                                    cell.fill = PatternFill("solid", fgColor="AED6F1")
                            except: pass
            except: pass

        all_df_list.append(df_sorted)

    if not all_df_list:
        wb.save(path)
        return path

    combined = pd.concat(all_df_list, ignore_index=True)

    # ── Top Performers tab ────────────────────────────────────────────────────
    top_picks = combined[combined["Tier"].isin(["S", "A"])].sort_values(
                    "Score", ascending=False).reset_index(drop=True)
    decent    = combined[combined["Tier"] == "B"].sort_values(
                    "Score", ascending=False).reset_index(drop=True)

    ws_top = wb.create_sheet(title="⭐ Top Performers", index=0)
    ws_top.append(["GOOFY SCREENER — PHASE 6 RESULTS"])
    ws_top.append([f"Run: {today}  |  Markets: US + ASX + JPX  |  "
                   f"Assets screened: {len(combined)}"])
    ws_top.append([])
    ws_top.append(["PHASE 6 ADDITIONS", "EXPLANATION"])
    for row in [
        ("Cluster",    f"Which correlated group (ρ>{CORR_THRESHOLD}) this asset belongs to. "
                       "'Unique' = no strong correlations with other TRADE assets."),
        ("Corr Risk",  "Low (Unique) / Medium (2–3 in cluster) / High (4+ in cluster). "
                       "High = concentrated bet even if assets look different."),
        ("Adj Size %", "Phase 5 size × (1 / √cluster_size). "
                       "Reduces correlated positions without zeroing any."),
    ]:
        ws_top.append(list(row))
    ws_top.append([])
    ws_top.append(["── S & A TIER ──"])
    if not top_picks.empty:
        cols = ["Market", "Asset", "Best Strategy", "Tier", "Score",
                "OUT Sharpe", "OUT Strat Ret %", "OUT Strat Max DD %",
                "Today's Verdict", "Kelly %", "Recommended Size %",
                "Cluster", "Corr Risk", "Adj Size %"]
        cols = [c for c in cols if c in top_picks.columns]
        ws_top.append(cols)
        for _, row in top_picks[cols].iterrows():
            ws_top.append([row[c] for c in cols])
    ws_top.append([])
    ws_top.append(["── B TIER ──"])
    if not decent.empty:
        cols = ["Market", "Asset", "Best Strategy", "Score",
                "OUT Sharpe", "OUT Strat Ret %", "Today's Verdict",
                "Cluster", "Corr Risk", "Adj Size %"]
        cols = [c for c in cols if c in decent.columns]
        ws_top.append(cols)
        for _, row in decent[cols].iterrows():
            ws_top.append([row[c] for c in cols])
    if EXCEL_FORMAT:
        try:
            ws_top["A1"].font = Font(bold=True, size=14, color="1C2833")
            ws_top["A2"].font = Font(italic=True, size=10, color="555555")
        except: pass

    # ── Today's Trade List ────────────────────────────────────────────────────
    trade_today = combined[
        (combined["Today's Verdict"] == "TRADE") &
        (combined["Tier"].isin(["S", "A", "B"]))
    ].sort_values("Adj Size %", ascending=False).reset_index(drop=True)

    ws_tt = wb.create_sheet(title="🟢 Today's Trade List", index=1)
    ws_tt.append(["GOOFY SCREENER — TODAY'S TRADE LIST (Phase 6)"])
    ws_tt.append([f"Run: {today}  |  Sorted by Adj Size % (Phase 6 correlation-adjusted)"])
    ws_tt.append([])

    if not trade_today.empty:
        cols = ["Market", "Asset", "Best Strategy", "Tier", "Score",
                "OUT Sharpe", "OUT Strat Ret %", "OUT Strat Max DD %",
                "Current Trend", "Today's Verdict",
                "Kelly %", "Recommended Size %",
                "Cluster", "Corr Risk", "Adj Size %"]
        cols = [c for c in cols if c in trade_today.columns]
        ws_tt.append(cols)
        for _, row in trade_today[cols].iterrows():
            ws_tt.append([row[c] for c in cols])
        _style_header(ws_tt, row_num=4)
        ws_tt.auto_filter.ref = f"A4:{get_column_letter(len(cols))}{ws_tt.max_row}"
        ws_tt.freeze_panes   = "A5"
        for col in ws_tt.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws_tt.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)
        if EXCEL_FORMAT:
            try:
                headers = cols
                for row_cells in ws_tt.iter_rows(min_row=5, max_row=ws_tt.max_row):
                    for cell in row_cells:
                        h = headers[cell.column - 1] if cell.column - 1 < len(headers) else ""
                        if h == "Corr Risk" and cell.value in CORR_RISK_COLORS:
                            cell.fill = PatternFill("solid",
                                fgColor=CORR_RISK_COLORS[cell.value])
                            cell.font = Font(bold=True, color="FFFFFF", size=10)
                        elif h == "Adj Size %" and cell.value is not None:
                            try:
                                v = float(cell.value)
                                if v >= 40:
                                    cell.fill = PatternFill("solid", fgColor="27AE60")
                                elif v >= 15:
                                    cell.fill = PatternFill("solid", fgColor="F39C12")
                                elif v > 0:
                                    cell.fill = PatternFill("solid", fgColor="AED6F1")
                            except: pass
            except: pass

    if EXCEL_FORMAT:
        try:
            ws_tt["A1"].font = Font(bold=True, size=14, color="27AE60")
            ws_tt["A2"].font = Font(italic=True, size=10, color="555555")
        except: pass

    # ── Correlation Matrix tab ────────────────────────────────────────────────
    corr_matrix = p6.get("corr_matrix", pd.DataFrame())
    ws_corr     = wb.create_sheet(title="🔗 Correlation Matrix")
    ws_corr.append(["PHASE 6 — PAIRWISE RETURN CORRELATIONS (TRADE-signal assets, test period)"])
    ws_corr.append([f"Cluster threshold: ρ ≥ {CORR_THRESHOLD}  |  "
                    f"Computed on 2021-present daily returns  |  "
                    f"Red = high correlation (concentrated risk), Blue = low (diversified)"])
    ws_corr.append([])

    if not corr_matrix.empty:
        tickers = list(corr_matrix.columns)
        header  = [""] + tickers
        ws_corr.append(header)
        data_start_row = ws_corr.max_row

        for ticker in tickers:
            row_vals = [ticker]
            for other in tickers:
                val = corr_matrix.loc[ticker, other]
                row_vals.append(round(float(val), 2) if pd.notna(val) else None)
            ws_corr.append(row_vals)

        # Conditional formatting: blue → white → red color scale
        if EXCEL_FORMAT:
            try:
                n     = len(tickers)
                c_end = get_column_letter(n + 1)
                r_end = data_start_row + n
                data_range = f"B{data_start_row}:{c_end}{r_end}"
                ws_corr.conditional_formatting.add(
                    data_range,
                    ColorScaleRule(
                        start_type="num",  start_value=-1, start_color="4FC3F7",  # blue
                        mid_type="num",    mid_value=0,    mid_color="FFFFFF",     # white
                        end_type="num",    end_value=1,    end_color="E53935",     # red
                    )
                )
                # Bold ticker labels
                for col_idx in range(2, n + 2):
                    ws_corr.cell(row=data_start_row, column=col_idx).font = Font(bold=True, size=9)
                for row_idx in range(data_start_row + 1, data_start_row + n + 1):
                    ws_corr.cell(row=row_idx, column=1).font = Font(bold=True, size=9)
                # Diagonal in grey
                for i in range(n):
                    cell = ws_corr.cell(row=data_start_row + 1 + i, column=2 + i)
                    cell.fill = PatternFill("solid", fgColor="CCCCCC")
                # Auto-width
                for col in ws_corr.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=0)
                    ws_corr.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 18)
            except: pass

        ws_corr.freeze_panes = f"B{data_start_row + 1}"

    if EXCEL_FORMAT:
        try:
            ws_corr["A1"].font = Font(bold=True, size=13, color="1C2833")
            ws_corr["A2"].font = Font(italic=True, size=10, color="555555")
        except: pass

    # ── Portfolio View tab ────────────────────────────────────────────────────
    port_m    = p6.get("port_metrics", {})
    adj_sizes = p6.get("adj_sizes", {})
    corr_risk = p6.get("corr_risk", {})
    cluster_sizes = p6.get("cluster_sizes", {})
    cluster_map   = p6.get("cluster_map", {})

    ws_pv = wb.create_sheet(title="📊 Portfolio View")
    ws_pv.append(["PHASE 6 — PORTFOLIO VIEW"])
    ws_pv.append([f"Run: {today}  |  TRADE-signal S/A/B assets only  |  "
                  f"Correlation threshold: ρ ≥ {CORR_THRESHOLD}"])
    ws_pv.append([])

    # Portfolio metrics summary
    ws_pv.append(["PORTFOLIO METRICS"])
    if port_m:
        ws_pv.append(["Portfolio Vol (ann.)",   f"{port_m.get('port_vol_ann', '—')}%"])
        ws_pv.append(["Diversification Ratio",  port_m.get("div_ratio", "—")])
        ws_pv.append(["Effective N (indep. bets)", port_m.get("effective_n", "—")])
        ws_pv.append(["Positions included",     port_m.get("n_positions", "—")])
    else:
        ws_pv.append(["Not enough TRADE-signal assets to compute portfolio metrics."])
    ws_pv.append([])

    # Final allocation table
    trade_sized = combined[
        (combined["Today's Verdict"] == "TRADE") &
        (combined["Tier"].isin(["S", "A", "B"])) &
        (combined["Asset"].isin(adj_sizes))
    ].copy()

    if not trade_sized.empty:
        trade_sized["Adj Size %"] = trade_sized["Asset"].map(adj_sizes)
        trade_sized = trade_sized.sort_values("Adj Size %", ascending=False).reset_index(drop=True)

        ws_pv.append(["FINAL ALLOCATION (sorted by Adj Size %)"])
        cols_pv = ["Market", "Asset", "Tier", "Score",
                   "OUT Sharpe", "OUT Strat Ret %",
                   "Kelly %", "Recommended Size %",
                   "Cluster", "Corr Risk", "Adj Size %"]
        cols_pv = [c for c in cols_pv if c in trade_sized.columns]
        pv_header_row = ws_pv.max_row + 1
        ws_pv.append(cols_pv)
        for _, row in trade_sized[cols_pv].iterrows():
            ws_pv.append([row[c] for c in cols_pv])

        _style_header(ws_pv, row_num=pv_header_row)
        ws_pv.auto_filter.ref = f"A{pv_header_row}:{get_column_letter(len(cols_pv))}{ws_pv.max_row}"
        ws_pv.freeze_panes    = f"A{pv_header_row + 1}"

        total_adj = sum(v for v in adj_sizes.values() if v)
        ws_pv.append([])
        ws_pv.append(["Total Adj Allocation", f"{total_adj:.1f}%"])

        if EXCEL_FORMAT:
            try:
                headers = cols_pv
                for row_cells in ws_pv.iter_rows(
                        min_row=pv_header_row + 1, max_row=ws_pv.max_row - 2):
                    for cell in row_cells:
                        h = headers[cell.column - 1] if cell.column - 1 < len(headers) else ""
                        if h == "Corr Risk" and cell.value in CORR_RISK_COLORS:
                            cell.fill = PatternFill("solid",
                                fgColor=CORR_RISK_COLORS[cell.value])
                            cell.font = Font(bold=True, color="FFFFFF", size=10)
                        elif h == "Adj Size %" and cell.value is not None:
                            try:
                                v = float(cell.value)
                                if v >= 40:
                                    cell.fill = PatternFill("solid", fgColor="27AE60")
                                elif v >= 15:
                                    cell.fill = PatternFill("solid", fgColor="F39C12")
                                elif v > 0:
                                    cell.fill = PatternFill("solid", fgColor="AED6F1")
                            except: pass
                for col in ws_pv.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=0)
                    ws_pv.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)
            except: pass

    if EXCEL_FORMAT:
        try:
            ws_pv["A1"].font = Font(bold=True, size=14, color="1C2833")
            ws_pv["A2"].font = Font(italic=True, size=10, color="555555")
        except: pass

    # ── Strategy Distribution tab ─────────────────────────────────────────────
    ws_dist = wb.create_sheet(title="📈 Strategy Distribution")
    ws_dist.append(["Strategy", "US Count", "ASX Count", "JPX Count", "Total"])
    for strat in STRATEGY_FNS.keys():
        row_data = [strat]
        total    = 0
        for market in ["US", "ASX", "JPX"]:
            c = (all_results.get(market, pd.DataFrame())
                 .get("Best Strategy", pd.Series()) == strat).sum() \
                if market in all_results and not all_results[market].empty else 0
            row_data.append(c); total += c
        row_data.append(total)
        ws_dist.append(row_data)
    _style_header(ws_dist)

    wb.save(path)
    return path


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Goofy Phase 6 — Portfolio construction with correlation analysis")
    parser.add_argument("--market", choices=["US", "ASX", "JPX", "ALL"],
                        default="ALL")
    args  = parser.parse_args()
    today = dt.datetime.now().strftime("%Y-%m-%d")

    loaded_gates = load_asset_gates()
    if loaded_gates:
        print(f"  [Phase 4] {sum(len(v) for v in loaded_gates.values())} "
              f"asset-specific gates across {len(loaded_gates)} assets")
    else:
        print("  [Phase 4] No asset_specific_gates.json — using theory defaults")

    markets_to_run = ["US", "ASX", "JPX"] if args.market == "ALL" else [args.market]

    print(f"""
╔══════════════════════════════════════════════════════════════════════╗
║  GOOFY SCREENER — PHASE 6  |  {today}                    ║
║  Regime-aware  +  Kelly/Vol sizing  +  Correlation portfolio layer   ║
║  Markets: {', '.join(markets_to_run):52} ║
║  Corr threshold: ρ ≥ {CORR_THRESHOLD}  |  Size adjustment: 1/√cluster_size  ║
╚══════════════════════════════════════════════════════════════════════╝""")

    # ── Step 1: Download ──────────────────────────────────────────────────────
    all_assets = []
    for m in markets_to_run:
        all_assets.extend(UNIVERSE_MAP[m])
    seen = set(); unique_assets = []
    for a in all_assets:
        if a not in seen:
            unique_assets.append(a); seen.add(a)

    print(f"\n[1/4] Downloading {len(unique_assets)} assets...\n")
    price_data = {}
    ohlc_data  = {}
    for asset in unique_assets:
        try:
            raw = yf.download(asset, start=TRAIN_START, end=TEST_END,
                              auto_adjust=True, progress=False)
            if not raw.empty and len(raw) >= MIN_ROWS:
                if isinstance(raw.columns, pd.MultiIndex):
                    raw.columns = raw.columns.get_level_values(0)
                close = raw["Close"].squeeze()
                if isinstance(close, pd.DataFrame):
                    close = close.iloc[:, 0]
                price_data[asset] = close
                if {"High", "Low", "Close"}.issubset(set(raw.columns)):
                    ohlc_data[asset] = raw[["High", "Low", "Close"]].copy()
                print(f"  ✓ {asset:14} | {len(raw)} rows")
            else:
                print(f"  ✗ {asset:14} | skipped ({len(raw)} rows)")
        except Exception as e:
            print(f"  ✗ {asset:14} | error: {e}")

    print(f"\n  → {len(price_data)} assets ready\n")

    # ── Step 2: Phase 5 screening (strategy + regime + sizing) ───────────────
    print("[2/4] Phase 5 screening (strategy + regime + sizing)...\n")
    all_results = {}
    for m in markets_to_run:
        all_results[m] = screen_market(m, UNIVERSE_MAP[m], price_data, ohlc_data)

    all_dfs = [df for df in all_results.values() if not df.empty]
    if not all_dfs:
        print("  No results."); return
    combined = pd.concat(all_dfs, ignore_index=True)

    # ── Step 3: Phase 6 portfolio construction ───────────────────────────────
    print("\n[3/4] Phase 6 — building correlation matrix + clusters + adjusted sizes...\n")
    p6 = run_portfolio_construction(combined, price_data)

    corr_m = p6.get("corr_matrix", pd.DataFrame())
    if not corr_m.empty:
        n_assets = len(corr_m)
        cluster_map   = p6["cluster_map"]
        cluster_sizes = p6["cluster_sizes"]
        n_unique  = sum(1 for a, n in cluster_sizes.items() if n == 1)
        n_grouped = len(cluster_sizes) - n_unique
        print(f"  Corr matrix:  {n_assets} × {n_assets} TRADE-signal assets")
        print(f"  Unique assets (no strong correlations): {n_unique}")
        print(f"  Grouped into clusters (ρ ≥ {CORR_THRESHOLD}):     {n_grouped}")

        pm = p6.get("port_metrics", {})
        if pm.get("port_vol_ann"):
            print(f"\n  Portfolio vol (ann.):       {pm['port_vol_ann']}%")
            print(f"  Diversification ratio:      {pm['div_ratio']}")
            print(f"  Effective independent bets: {pm['effective_n']}")
    else:
        print("  Not enough TRADE-signal assets for correlation analysis.")

    # Add Phase 6 columns to results
    all_results = attach_phase6_columns(all_results, p6)
    combined    = pd.concat([df for df in all_results.values() if not df.empty],
                            ignore_index=True)

    # ── Step 4: Save + console summary ───────────────────────────────────────
    print(f"\n[4/4] Saving Phase 6 report...\n")

    print(f"\n{'═'*72}")
    print(f"  GOOFY SCREENER — PHASE 6 SUMMARY  |  {today}")
    print(f"{'═'*72}")
    for tier in ["S", "A", "B", "Skip"]:
        icon  = {"S": "⭐", "A": "✅", "B": "🔵", "Skip": "⬜"}.get(tier, "")
        count = len(combined[combined["Tier"] == tier])
        print(f"    {icon} {tier:5}: {count:3}")

    adj_sizes = p6.get("adj_sizes", {})
    trade_list = combined[
        (combined["Today's Verdict"] == "TRADE") &
        (combined["Tier"].isin(["S", "A", "B"]))
    ].sort_values("Score", ascending=False)

    if not trade_list.empty:
        print(f"\n  {'Asset':14} {'Strategy':18} {'Tier':4} "
              f"{'Phase5%':8} {'Cluster':10} {'CorrRisk':9} {'AdjSize%':8}")
        print(f"  {'─'*75}")
        for _, r in trade_list.iterrows():
            p5  = f"{r['Recommended Size %']:.0f}%" if r.get('Recommended Size %') else "—"
            adj = f"{adj_sizes.get(r['Asset'], 0):.0f}%"
            clus = r.get("Cluster", "—")
            crisk = r.get("Corr Risk", "—")
            print(f"  {r['Asset']:14} {r['Best Strategy']:18} {r['Tier']:4} "
                  f"{p5:>8} {clus:>10} {crisk:>9} {adj:>8}")

    xlsx_path = write_excel_phase6(all_results, today, p6)
    print(f"\n  Report saved → {os.path.basename(xlsx_path)}")
    print(f"  Full path:   {xlsx_path}")
    print(f"\n{'═'*72}\n")


if __name__ == "__main__":
    main()
