"""
╔══════════════════════════════════════════════════════════════════════╗
║           GOOFY SCREENER — PHASE 5                                   ║
║           Position Sizing: Kelly Criterion + Volatility Scaling      ║
║                                                                      ║
║   Inherits everything from Phase 4 (regime-aware verdict) — adds    ║
║   a "how much?" answer on top of "should I trade today?".           ║
║                                                                      ║
║   Two methods, combined:                                             ║
║     Half-Kelly  — sizes based on win rate & avg win/loss            ║
║     Vol Scaling — adjusts so every asset targets 15% annual vol     ║
║                                                                      ║
║   Three new columns per asset:                                       ║
║       Kelly %           — half-Kelly fraction                        ║
║       Vol Scalar        — volatility scaling multiplier              ║
║       Recommended Size% — Kelly × Vol Scalar, capped at 100%        ║
║                                                                      ║
║   New Excel tab:                                                     ║
║       📐 Position Sizing — all TRADE-verdict assets sized up         ║
║                                                                      ║
║   Run:  python3 goofy_screener_phase5.py                            ║
║   Args: --market US | ASX | JPX | ALL (default: ALL)               ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import yfinance as yf
import numpy as np
import pandas as pd
import datetime as dt
import os, sys, warnings, argparse
warnings.filterwarnings("ignore")

# Phase 4 regime layer
from regime_detector import (
    label_regimes,
    get_allowed_regimes,
    load_asset_gates,
    ASSET_SPECIFIC_GATES,
)

# Phase 5 position sizing layer
from position_sizer import (
    compute_trade_stats,
    recommend_size,
)

# ── Optional: rich Excel formatting ───────────────────────────────────────────
try:
    from openpyxl import load_workbook
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
    from openpyxl.utils import get_column_letter
    EXCEL_FORMAT = True
except ImportError:
    EXCEL_FORMAT = False
    print("  [INFO] openpyxl not found — plain Excel only. Run: pip install openpyxl")

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIG
# ══════════════════════════════════════════════════════════════════════════════
TRAIN_START      = dt.datetime(2016, 1, 1)
TRAIN_END        = dt.datetime(2021, 1, 1)
TEST_END         = dt.datetime.now()
MIN_ROWS         = 400
PERIODS_PER_YEAR = 252

TARGET_VOL       = 0.15   # 15% annual — volatility scaling target
KELLY_FRACTION   = 0.5    # half-Kelly
VOL_LOOKBACK     = 21     # ~1 month of trading days for recent vol estimate

TIER_S = {"sharpe": 0.8,  "ret": 30,  "max_dd": -20}
TIER_A = {"sharpe": 0.4,  "ret": 10,  "max_dd": -35}
TIER_B = {"sharpe": 0.1,  "ret": -10, "max_dd": -50}

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "screener_output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

LONG_TO_SHORT = {
    "MA Crossover":    "MA",
    "RSI":             "RSI",
    "Bollinger Bands": "BB",
    "MACD":            "MACD",
    "Mean Reversion":  "MeanReversion",
}

# ══════════════════════════════════════════════════════════════════════════════
#  STOCK UNIVERSES  (unchanged from Phase 4)
# ══════════════════════════════════════════════════════════════════════════════

US_UNIVERSE = [
    "AAPL", "MSFT", "NVDA", "GOOGL", "META", "AMZN", "TSLA", "AMD",
    "JPM", "BAC", "GS", "MS", "V", "MA", "BRK-B",
    "JNJ", "UNH", "LLY", "PFE", "ABBV", "MRK",
    "XOM", "CVX", "COP",
    "PG", "KO", "PEP", "WMT", "COST", "MCD",
    "CAT", "DE", "BA", "HON", "RTX",
    "SPY", "QQQ", "GLD", "TLT", "IWM",
]

ASX_UNIVERSE = [
    "CBA.AX", "WBC.AX", "ANZ.AX", "NAB.AX",
    "BHP.AX", "RIO.AX", "FMG.AX", "S32.AX",
    "WDS.AX", "STO.AX", "BPT.AX",
    "CSL.AX", "RMD.AX", "COH.AX", "SHL.AX",
    "WES.AX", "WOW.AX", "COL.AX", "JBH.AX",
    "XRO.AX", "WTC.AX", "ALU.AX",
    "TCL.AX", "APA.AX", "SKI.AX",
    "GMG.AX", "SCG.AX", "GPT.AX",
    "IOZ.AX", "STW.AX", "VAS.AX",
]

JPX_UNIVERSE = [
    "7203.T", "7267.T", "7261.T", "7272.T", "7269.T",
    "6758.T", "6501.T", "6954.T", "6902.T", "6861.T",
    "6762.T", "8035.T", "6857.T", "6723.T", "6594.T",
    "4063.T", "4523.T",
    "9432.T", "9433.T", "9434.T",
    "9984.T", "4689.T", "6098.T", "4385.T",
    "7974.T", "9684.T", "7832.T",
    "8306.T", "8316.T", "8411.T",
    "8750.T", "8725.T",
    "3382.T", "8267.T", "4661.T", "2914.T",
    "6301.T", "6326.T", "7011.T",
    "4502.T", "4519.T",
    "8001.T", "8002.T", "8058.T",
    "9022.T", "9020.T",
    "1321.T", "1306.T",
]

UNIVERSE_MAP = {
    "US":  US_UNIVERSE,
    "ASX": ASX_UNIVERSE,
    "JPX": JPX_UNIVERSE,
}


# ══════════════════════════════════════════════════════════════════════════════
#  METRICS CALCULATOR
# ══════════════════════════════════════════════════════════════════════════════

def compute_metrics(price: pd.Series, position: pd.Series) -> dict:
    df = pd.DataFrame({"price": price})
    df["pos"]       = position.reindex(df.index).fillna(0)
    df["log_ret"]   = np.log(df["price"] / df["price"].shift(1))
    df["strat_ret"] = df["pos"] * df["log_ret"]
    df["cum"]       = np.exp(df["strat_ret"].cumsum())

    lr = df["strat_ret"].dropna()
    if len(lr) < 10:
        return {"Sharpe": np.nan, "TotalRet": np.nan, "MaxDD": np.nan, "WinRate": np.nan}

    ann    = np.exp(lr.mean() * PERIODS_PER_YEAR) - 1
    vol    = lr.std() * np.sqrt(PERIODS_PER_YEAR)
    sharpe = ann / vol if vol != 0 else np.nan
    cum    = df["cum"].dropna()
    mdd    = ((cum - cum.cummax()) / cum.cummax()).min() if len(cum) > 0 else np.nan
    total  = cum.iloc[-1] - 1 if len(cum) > 0 else np.nan
    wins   = (lr > 0).sum() / max(len(lr), 1)

    return {
        "Sharpe":   round(float(sharpe), 3) if not np.isnan(sharpe) else np.nan,
        "TotalRet": round(float(total) * 100, 2) if not np.isnan(total) else np.nan,
        "MaxDD":    round(float(mdd) * 100, 2) if not np.isnan(mdd) else np.nan,
        "WinRate":  round(float(wins) * 100, 1) if not np.isnan(wins) else np.nan,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  STRATEGY FUNCTIONS  (unchanged from Phase 4)
# ══════════════════════════════════════════════════════════════════════════════

def strategy_ma(p, short=20, long=50):
    return (p.rolling(short).mean() > p.rolling(long).mean()).astype(int).shift(1)

def strategy_rsi(p, period=14, oversold=30, overbought=70):
    d   = p.diff()
    rsi = 100 - (100 / (1 + d.clip(lower=0).rolling(period).mean() /
                            (-d.clip(upper=0)).rolling(period).mean()))
    sig  = np.zeros(len(p)); hold = False
    for i, r in enumerate(rsi):
        if np.isnan(r): sig[i] = 0
        elif not hold and r < oversold:  hold = True;  sig[i] = 1
        elif hold and r > overbought:    hold = False; sig[i] = 0
        else:                            sig[i] = 1 if hold else 0
    return pd.Series(sig, index=p.index).shift(1)

def strategy_bb(p, window=20, num_std=2.0):
    mid = p.rolling(window).mean()
    low = mid - num_std * p.rolling(window).std()
    sig = np.zeros(len(p)); hold = False
    for i in range(len(p)):
        pi, m, l = p.iloc[i], mid.iloc[i], low.iloc[i]
        if np.isnan(l): sig[i] = 0
        elif not hold and pi <= l: hold = True;  sig[i] = 1
        elif hold and pi >= m:     hold = False; sig[i] = 0
        else:                      sig[i] = 1 if hold else 0
    return pd.Series(sig, index=p.index).shift(1)

def strategy_macd(p, fast=12, slow=26, signal_p=9):
    macd = p.ewm(span=fast).mean() - p.ewm(span=slow).mean()
    return (macd > macd.ewm(span=signal_p).mean()).astype(int).shift(1)

def strategy_mr(p, window=20, threshold=1.5):
    z   = (p - p.rolling(window).mean()) / p.rolling(window).std()
    sig = np.zeros(len(p)); hold = False
    for i, zi in enumerate(z):
        if np.isnan(zi): sig[i] = 0
        elif not hold and zi < -threshold: hold = True;  sig[i] = 1
        elif hold and zi >= 0:             hold = False; sig[i] = 0
        else:                              sig[i] = 1 if hold else 0
    return pd.Series(sig, index=p.index).shift(1)

STRATEGY_FNS = {
    "MA Crossover":    strategy_ma,
    "RSI":             strategy_rsi,
    "Bollinger Bands": strategy_bb,
    "MACD":            strategy_macd,
    "Mean Reversion":  strategy_mr,
}

STRATEGY_GRIDS = {
    "MA Crossover":    [{"short": s, "long": l}
                        for s in [10, 20, 50] for l in [50, 100, 200] if s < l],
    "RSI":             [{"period": p, "oversold": os, "overbought": ob}
                        for p in [10, 14, 21] for os, ob in [(25,65),(30,70),(35,75)]],
    "Bollinger Bands": [{"window": w, "num_std": s}
                        for w in [10, 20, 30] for s in [1.5, 2.0, 2.5]],
    "MACD":            [{"fast": f, "slow": s, "signal_p": sg}
                        for f in [8, 12] for s in [21, 26] for sg in [7, 9] if f < s],
    "Mean Reversion":  [{"window": w, "threshold": t}
                        for w in [10, 20, 40] for t in [1.0, 1.5, 2.0]],
}


# ══════════════════════════════════════════════════════════════════════════════
#  SCORING & TIERING  (unchanged from Phase 4)
# ══════════════════════════════════════════════════════════════════════════════

def score_asset(row: dict):
    sharpe  = row.get("OUT Sharpe", np.nan)
    ret     = row.get("OUT Strat Ret %", np.nan)
    mdd     = row.get("OUT Strat Max DD %", np.nan)
    dd_save = row.get("DD Saved %", np.nan)

    if any(pd.isna(v) for v in [sharpe, ret, mdd]):
        return ("Skip", 0.0)

    s_pts  = min(max(sharpe / 1.5, 0), 1) * 40
    r_norm = max(ret + 30, 0) / 130
    r_pts  = min(r_norm, 1) * 25
    d_pts  = min(max((60 + mdd) / 60, 0), 1) * 20
    ds_pts = min(max((dd_save + 20) / 60, 0), 1) * 15 if not pd.isna(dd_save) else 0

    score = round(s_pts + r_pts + d_pts + ds_pts, 1)

    if sharpe >= TIER_S["sharpe"] and ret >= TIER_S["ret"] and mdd >= TIER_S["max_dd"]:
        tier = "S"
    elif sharpe >= TIER_A["sharpe"] and ret >= TIER_A["ret"] and mdd >= TIER_A["max_dd"]:
        tier = "A"
    elif sharpe >= TIER_B["sharpe"] and mdd >= TIER_B["max_dd"]:
        tier = "B"
    else:
        tier = "Skip"

    return (tier, score)


# ══════════════════════════════════════════════════════════════════════════════
#  REGIME VERDICT  (Phase 4 layer, unchanged)
# ══════════════════════════════════════════════════════════════════════════════

def compute_today_verdict(asset: str, best_strategy_long: str,
                          ohlc: pd.DataFrame) -> dict:
    blank = {
        "Current Trend": "—", "Current Vol": "—",
        "Allowed Regimes": "—", "Today's Verdict": "—",
    }
    if ohlc is None or ohlc.empty:
        return blank

    needed = {"High", "Low", "Close"}
    if not needed.issubset(set(ohlc.columns)):
        return blank

    try:
        labelled  = label_regimes(ohlc)
        trend_now = labelled["Trend"].dropna()
        vol_now   = labelled["Vol"].dropna()
        if trend_now.empty:
            return blank
        cur_trend = trend_now.iloc[-1]
        cur_vol   = vol_now.iloc[-1] if not vol_now.empty else "—"
    except Exception:
        return blank

    short = LONG_TO_SHORT.get(best_strategy_long)
    if short is None:
        return {"Current Trend": cur_trend, "Current Vol": cur_vol,
                "Allowed Regimes": "—", "Today's Verdict": "—"}

    allowed     = get_allowed_regimes(asset, short)
    allowed_str = ", ".join(sorted(allowed)) if allowed else "(no gate)"

    if not allowed:
        verdict = "TRADE"
    else:
        verdict = "TRADE" if cur_trend in allowed else "STAND DOWN"

    return {
        "Current Trend":   cur_trend,
        "Current Vol":     cur_vol,
        "Allowed Regimes": allowed_str,
        "Today's Verdict": verdict,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  MARKET SCREENER — single market pass
# ══════════════════════════════════════════════════════════════════════════════

def screen_market(market_name: str, assets: list, price_data: dict,
                  ohlc_data: dict) -> pd.DataFrame:
    today   = dt.datetime.now().strftime("%Y-%m-%d")
    results = []

    print(f"\n  ── Screening {market_name} "
          f"({len([a for a in assets if a in price_data])} valid assets) ──")

    for asset in assets:
        if asset not in price_data:
            continue

        full  = price_data[asset]
        train = full[full.index < TRAIN_END]
        test  = full[full.index >= TRAIN_END]

        if len(train) < 100 or len(test) < 50:
            continue

        # ── Best strategy search (train period) ──────────────────────────────
        best_sharpe = -999
        best_strat  = None
        best_params = None

        for strat_name, fn in STRATEGY_FNS.items():
            for params in STRATEGY_GRIDS[strat_name]:
                try:
                    pos = fn(train, **params)
                    m   = compute_metrics(train, pos)
                    if not np.isnan(m["Sharpe"]) and m["Sharpe"] > best_sharpe:
                        best_sharpe = m["Sharpe"]
                        best_strat  = strat_name
                        best_params = params
                except Exception:
                    continue

        if best_strat is None:
            continue

        # ── Out-of-sample (test period) ──────────────────────────────────────
        pos_test = STRATEGY_FNS[best_strat](test, **best_params)
        tm       = compute_metrics(test, pos_test)

        # ── B&H benchmark ────────────────────────────────────────────────────
        log      = np.log(test / test.shift(1)).dropna()
        bah_ret  = (np.exp(log.cumsum()).iloc[-1] - 1) * 100 if len(log) > 0 else np.nan
        bah_cum  = np.exp(log.cumsum()) if len(log) > 0 else pd.Series(dtype=float)
        bah_mdd  = ((bah_cum - bah_cum.cummax()) / bah_cum.cummax()).min() * 100 \
                   if len(bah_cum) > 0 else np.nan
        dd_saved = round(tm["MaxDD"] - float(bah_mdd), 2) \
                   if not np.isnan(bah_mdd) else np.nan

        # ── Phase 4: regime verdict ───────────────────────────────────────────
        verdict = compute_today_verdict(asset, best_strat, ohlc_data.get(asset))

        # ── Phase 5: position sizing ──────────────────────────────────────────
        trade_stats    = compute_trade_stats(test, pos_test)
        recent_returns = test.pct_change().dropna()
        sizing         = recommend_size(
            trade_stats,
            recent_returns,
            target_vol=TARGET_VOL,
            kelly_fraction=KELLY_FRACTION,
        )

        row = {
            "Market":             market_name,
            "Asset":              asset,
            "Best Strategy":      best_strat,
            "Best Params":        str(best_params),
            "Train Sharpe":       round(best_sharpe, 3),
            "OUT Sharpe":         tm["Sharpe"],
            "OUT Win Rate %":     tm["WinRate"],
            "OUT Strat Ret %":    tm["TotalRet"],
            "OUT B&H Ret %":      round(float(bah_ret), 2) if not np.isnan(bah_ret) else np.nan,
            "OUT Strat Max DD %": tm["MaxDD"],
            "OUT B&H Max DD %":   round(float(bah_mdd), 2) if not np.isnan(bah_mdd) else np.nan,
            "DD Saved %":         dd_saved,
            "Beats B&H":          (tm["TotalRet"] or 0) > (bah_ret or 0),
            # Phase 4 regime columns
            "Current Trend":      verdict["Current Trend"],
            "Current Vol":        verdict["Current Vol"],
            "Allowed Regimes":    verdict["Allowed Regimes"],
            "Today's Verdict":    verdict["Today's Verdict"],
            # Phase 5 position sizing columns
            "N Trades":           trade_stats["n_trades"] if trade_stats else None,
            "Trade Win Rate %":   round(trade_stats["win_rate"] * 100, 1) if trade_stats else None,
            "Avg Win %":          round(trade_stats["avg_win"] * 100, 2) if trade_stats else None,
            "Avg Loss %":         round(trade_stats["avg_loss"] * 100, 2) if trade_stats else None,
            "Kelly %":            sizing["Kelly %"],
            "Vol Scalar":         sizing["Vol Scalar"],
            "Recommended Size %": sizing["Recommended Size %"],
            "Run Date":           today,
        }

        tier, score = score_asset(row)
        row["Tier"]  = tier
        row["Score"] = score
        results.append(row)

        # Console output
        tier_icon = {"S": "⭐", "A": "✅", "B": "🔵", "Skip": "⬜"}.get(tier, "")
        beats_icon = "✓" if row["Beats B&H"] else "✗"
        dd_str    = f"{dd_saved:+.1f}%" if not np.isnan(dd_saved) else "N/A"
        v_icon    = {"TRADE": "🟢", "STAND DOWN": "🔴", "—": "  "}.get(
                        verdict["Today's Verdict"], "  ")
        size_str  = (f"{sizing['Recommended Size %']:.0f}%"
                     if sizing["Recommended Size %"] is not None else "—")
        v_verdict = verdict["Today's Verdict"]
        print(f"    {tier_icon} [{tier}] {asset:14} → {best_strat:18} | "
              f"Sharpe: {tm['Sharpe']:5.2f} | Ret: {(tm['TotalRet'] or 0):6.0f}% | "
              f"DD: {(tm['MaxDD'] or 0):5.1f}% | DDsaved: {dd_str:>8} | B&H {beats_icon} | "
              f"Score: {score:.0f}/100 | {v_icon} {verdict['Current Trend']:8} → "
              f"{v_verdict:10} | Size: {size_str}")

    return pd.DataFrame(results)


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL COLOURS
# ══════════════════════════════════════════════════════════════════════════════

TIER_COLORS = {
    "S":    "FFD700", "A": "90EE90", "B": "87CEEB", "Skip": "D3D3D3",
}
STRAT_COLORS = {
    "MA Crossover": "AED6F1", "RSI": "FAD7A0", "Bollinger Bands": "A9DFBF",
    "MACD": "D7BDE2", "Mean Reversion": "F1948A",
}
TREND_COLORS  = {"Bull": "D5F5E3", "Sideways": "FCF3CF", "Bear": "FADBD8"}
VERDICT_COLORS = {"TRADE": "27AE60", "STAND DOWN": "C0392B"}


def _style_header(ws, row_num=1, bg="1C2833"):
    if not EXCEL_FORMAT:
        return
    try:
        THIN = Side(style="thin", color="CCCCCC")
        BRD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        for cell in ws[row_num]:
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = BRD
        ws.row_dimensions[row_num].height = 28
    except Exception:
        pass


def apply_sheet_formatting(ws, df):
    if not EXCEL_FORMAT:
        return
    try:
        THIN = Side(style="thin", color="CCCCCC")
        BRD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        GRN  = PatternFill("solid", fgColor="D5F5E3")
        RED  = PatternFill("solid", fgColor="FADBD8")

        headers = [c.value for c in ws[1]]
        col_idx = {h: i+1 for i, h in enumerate(headers)}

        for cell in ws[1]:
            cell.fill      = PatternFill("solid", fgColor="1C2833")
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = BRD
        ws.row_dimensions[1].height = 30

        for row in ws.iter_rows(min_row=2):
            strat_val   = row[col_idx["Best Strategy"]-1].value   if "Best Strategy"   in col_idx else None
            tier_val    = row[col_idx["Tier"]-1].value            if "Tier"            in col_idx else None
            beats_val   = row[col_idx["Beats B&H"]-1].value       if "Beats B&H"       in col_idx else None
            trend_val   = row[col_idx["Current Trend"]-1].value   if "Current Trend"   in col_idx else None
            verdict_val = row[col_idx["Today's Verdict"]-1].value if "Today's Verdict" in col_idx else None

            for cell in row:
                cell.border    = BRD
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font      = Font(size=10)
                h = headers[cell.column - 1]

                if h == "Best Strategy" and strat_val:
                    cell.fill = PatternFill("solid", fgColor=STRAT_COLORS.get(strat_val, "FFFFFF"))
                    cell.font = Font(bold=True, size=10)
                elif h == "Tier" and tier_val:
                    cell.fill = PatternFill("solid", fgColor=TIER_COLORS.get(tier_val, "FFFFFF"))
                    cell.font = Font(bold=True, size=10)
                elif h == "Score":
                    try:
                        v = float(cell.value or 0)
                        if v >= 70:   cell.fill = PatternFill("solid", fgColor="27AE60")
                        elif v >= 50: cell.fill = PatternFill("solid", fgColor="F39C12")
                        elif v >= 30: cell.fill = PatternFill("solid", fgColor="AED6F1")
                        else:         cell.fill = PatternFill("solid", fgColor="E8E8E8")
                    except: pass
                elif h == "OUT Sharpe":
                    try:
                        v = float(cell.value or 0)
                        if v >= 0.8: cell.fill = GRN
                        elif v < 0:  cell.fill = RED
                    except: pass
                elif h == "OUT Strat Ret %":
                    try:
                        v = float(cell.value or 0)
                        if v >= 30:   cell.fill = GRN
                        elif v < -10: cell.fill = RED
                    except: pass
                elif h in ("OUT Strat Max DD %", "OUT B&H Max DD %"):
                    try:
                        v = float(cell.value or 0)
                        if v >= -20:  cell.fill = GRN
                        elif v < -40: cell.fill = RED
                    except: pass
                elif h == "DD Saved %":
                    try:
                        v = float(cell.value or 0)
                        if v >= 10: cell.fill = GRN
                        elif v < 0: cell.fill = RED
                    except: pass
                elif h == "Beats B&H":
                    cell.fill = GRN if beats_val else RED
                elif h == "Current Trend" and trend_val in TREND_COLORS:
                    cell.fill = PatternFill("solid", fgColor=TREND_COLORS[trend_val])
                    cell.font = Font(bold=True, size=10)
                elif h == "Today's Verdict" and verdict_val in VERDICT_COLORS:
                    cell.fill = PatternFill("solid", fgColor=VERDICT_COLORS[verdict_val])
                    cell.font = Font(bold=True, color="FFFFFF", size=10)
                elif h == "Recommended Size %" and cell.value is not None:
                    try:
                        v = float(cell.value)
                        if v >= 60:   cell.fill = PatternFill("solid", fgColor="27AE60")
                        elif v >= 30: cell.fill = PatternFill("solid", fgColor="F39C12")
                        elif v > 0:   cell.fill = PatternFill("solid", fgColor="AED6F1")
                        else:         cell.fill = RED
                    except: pass
                elif h in ("Kelly %", "Vol Scalar"):
                    cell.font = Font(italic=True, size=10)

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 32)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    except Exception as e:
        print(f"  [WARN] Formatting failed: {e}")


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL REPORTER
# ══════════════════════════════════════════════════════════════════════════════

def write_excel_report(all_results: dict, today: str) -> str:
    fname = f"Goofy_Phase5_{today}.xlsx"
    path  = os.path.join(OUTPUT_DIR, fname)

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    all_df_list  = []
    market_flags = {"US": "🇺🇸", "ASX": "🇦🇺", "JPX": "🇯🇵"}
    HIDE_COLS    = {"Best Params"}   # hide technical noise

    # ── Per-market tabs ───────────────────────────────────────────────────────
    for market_name, df in all_results.items():
        if df.empty:
            continue
        df_sorted = df.sort_values("Score", ascending=False).reset_index(drop=True)
        flag = market_flags.get(market_name, "")
        ws   = wb.create_sheet(title=f"{flag} {market_name}")

        cols_to_show = [c for c in df_sorted.columns if c not in HIDE_COLS]
        ws.append(cols_to_show)
        for _, row in df_sorted[cols_to_show].iterrows():
            ws.append([row[c] for c in cols_to_show])

        apply_sheet_formatting(ws, df_sorted[cols_to_show])
        all_df_list.append(df_sorted)

    if not all_df_list:
        wb.save(path)
        return path

    combined = pd.concat(all_df_list, ignore_index=True)

    # ── Top Performers tab ────────────────────────────────────────────────────
    top_picks = combined[combined["Tier"].isin(["S", "A"])].sort_values(
                    "Score", ascending=False).reset_index(drop=True)
    decent    = combined[combined["Tier"] == "B"].sort_values(
                    "Score", ascending=False).reset_index(drop=True)

    ws_top = wb.create_sheet(title="⭐ Top Performers", index=0)
    ws_top.append(["GOOFY SCREENER — PHASE 5 RESULTS"])
    ws_top.append([f"Run: {today}  |  Markets: US + ASX + JPX  |  "
                   f"Assets screened: {len(combined)}"])
    ws_top.append([])

    ws_top.append(["TIER", "CRITERIA"])
    for row in [("S (⭐ Excellent)", "Sharpe ≥ 0.8, Return ≥ 30%, Max DD ≥ -20%"),
                ("A (✅ Good)",      "Sharpe ≥ 0.4, Return ≥ 10%, Max DD ≥ -35%"),
                ("B (🔵 Decent)",   "Sharpe ≥ 0.1, Max DD ≥ -50%"),
                ("Skip (⬜)",        "Below all thresholds")]:
        ws_top.append(list(row))
    ws_top.append([])

    ws_top.append(["POSITION SIZING (Phase 5)", "EXPLANATION"])
    for row in [
        ("Kelly %",            "Half-Kelly fraction from win rate & avg win/loss"),
        ("Vol Scalar",         "Multiplier to target 15% annualised vol (>1 = calm asset, <1 = wild)"),
        ("Recommended Size %", "Kelly % × Vol Scalar, capped at 100% — how much capital to deploy"),
        ("—",                  "Not enough trades (<5) to estimate Kelly reliably"),
    ]:
        ws_top.append(list(row))
    ws_top.append([])

    ws_top.append(["TIER BREAKDOWN"])
    for tier in ["S", "A", "B", "Skip"]:
        ws_top.append([tier, len(combined[combined["Tier"] == tier])])
    ws_top.append([])

    ws_top.append(["── S & A TIER (Best Opportunities) ──"])
    if not top_picks.empty:
        cols = ["Market", "Asset", "Best Strategy", "Tier", "Score",
                "OUT Sharpe", "OUT Strat Ret %", "OUT B&H Ret %", "OUT Strat Max DD %",
                "DD Saved %", "Beats B&H", "Current Trend", "Today's Verdict",
                "Kelly %", "Vol Scalar", "Recommended Size %"]
        ws_top.append(cols)
        for _, row in top_picks[cols].iterrows():
            ws_top.append([row[c] for c in cols])
    else:
        ws_top.append(["No S or A tier assets found."])

    ws_top.append([])
    ws_top.append(["── B TIER (Decent — worth watching) ──"])
    if not decent.empty:
        cols = ["Market", "Asset", "Best Strategy", "Score",
                "OUT Sharpe", "OUT Strat Ret %", "OUT Strat Max DD %", "DD Saved %",
                "Current Trend", "Today's Verdict", "Recommended Size %"]
        ws_top.append(cols)
        for _, row in decent[cols].iterrows():
            ws_top.append([row[c] for c in cols])
    else:
        ws_top.append(["No B tier assets found."])

    if EXCEL_FORMAT:
        try:
            ws_top["A1"].font = Font(bold=True, size=14, color="1C2833")
            ws_top["A2"].font = Font(italic=True, size=10, color="555555")
        except: pass

    # ── Today's Trade List ────────────────────────────────────────────────────
    trade_today = combined[
        (combined["Today's Verdict"] == "TRADE") &
        (combined["Tier"].isin(["S", "A", "B"]))
    ].sort_values("Score", ascending=False).reset_index(drop=True)

    ws_tt = wb.create_sheet(title="🟢 Today's Trade List", index=1)
    ws_tt.append(["GOOFY SCREENER — TODAY'S TRADE LIST (Phase 5)"])
    ws_tt.append([f"Run: {today}  |  TRADE-verdict S/A/B tier assets  |  "
                  f"Includes Phase 5 position sizing"])
    ws_tt.append([])

    if not trade_today.empty:
        cols = ["Market", "Asset", "Best Strategy", "Tier", "Score",
                "OUT Sharpe", "OUT Strat Ret %", "OUT Strat Max DD %",
                "Current Trend", "Current Vol", "Today's Verdict",
                "N Trades", "Trade Win Rate %", "Avg Win %", "Avg Loss %",
                "Kelly %", "Vol Scalar", "Recommended Size %"]
        ws_tt.append(cols)
        for _, row in trade_today[cols].iterrows():
            ws_tt.append([row[c] for c in cols])
        _style_header(ws_tt, row_num=4)
        ws_tt.auto_filter.ref = f"A4:{get_column_letter(len(cols))}{ws_tt.max_row}"
        ws_tt.freeze_panes = "A5"

        for col in ws_tt.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws_tt.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)
    else:
        ws_tt.append(["No TRADE-verdict S/A/B assets today."])

    if EXCEL_FORMAT:
        try:
            ws_tt["A1"].font = Font(bold=True, size=14, color="27AE60")
            ws_tt["A2"].font = Font(italic=True, size=10, color="555555")
        except: pass

    # ── Position Sizing tab (Phase 5 NEW) ────────────────────────────────────
    ws_ps = wb.create_sheet(title="📐 Position Sizing")
    ws_ps.append(["PHASE 5 — POSITION SIZING BREAKDOWN"])
    ws_ps.append([f"Run: {today}  |  All screened assets  |  "
                  f"Target vol: {TARGET_VOL*100:.0f}%  |  Kelly fraction: "
                  f"{KELLY_FRACTION*100:.0f}% (half-Kelly)"])
    ws_ps.append([])

    ws_ps.append(["HOW TO READ THIS TABLE"])
    for row in [
        ("Kelly %",            "Based on OOS trade-by-trade stats. "
                               "Full Kelly × 0.5 = half-Kelly. Negative edge → 0%."),
        ("Vol Scalar",         f"target_vol ({TARGET_VOL*100:.0f}%) ÷ recent_vol (21-day). "
                               "Asset is calm → scalar > 1. Asset is wild → scalar < 1."),
        ("Recommended Size %", "Kelly % × Vol Scalar, capped at 100%. "
                               "This is the % of your capital to deploy for this trade."),
        ("N Trades",           "Fewer trades = less reliable Kelly estimate. "
                               "Use caution below 10 trades."),
    ]:
        ws_ps.append(list(row))
    ws_ps.append([])

    ws_ps.append(["── All assets with sizing data (sorted by Score) ──"])
    sized = combined[combined["Kelly %"].notna()].sort_values(
                "Score", ascending=False).reset_index(drop=True)
    no_size = combined[combined["Kelly %"].isna()].sort_values(
                "Score", ascending=False).reset_index(drop=True)

    if not sized.empty:
        cols = ["Market", "Asset", "Tier", "Score", "Best Strategy",
                "N Trades", "Trade Win Rate %", "Avg Win %", "Avg Loss %",
                "Kelly %", "Vol Scalar", "Recommended Size %",
                "Today's Verdict", "Current Trend"]
        ws_ps.append(cols)
        for _, row in sized[cols].iterrows():
            ws_ps.append([row[c] for c in cols])
        _style_header(ws_ps, row_num=8)
        ws_ps.auto_filter.ref = f"A8:{get_column_letter(len(cols))}{ws_ps.max_row}"
        ws_ps.freeze_panes = "A9"

    ws_ps.append([])
    ws_ps.append(["── Assets with insufficient trade data (Kelly = —) ──"])
    if not no_size.empty:
        cols2 = ["Market", "Asset", "Tier", "Score", "Best Strategy",
                 "N Trades", "Today's Verdict"]
        ws_ps.append(cols2)
        for _, row in no_size[cols2].iterrows():
            ws_ps.append([row[c] for c in cols2])

    if EXCEL_FORMAT:
        try:
            ws_ps["A1"].font = Font(bold=True, size=14, color="1C2833")
            ws_ps["A2"].font = Font(italic=True, size=10, color="555555")
            for col in ws_ps.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                ws_ps.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 36)
        except: pass

    # ── Strategy Distribution tab ─────────────────────────────────────────────
    ws_dist = wb.create_sheet(title="📊 Strategy Distribution")
    ws_dist.append(["Strategy", "US Count", "ASX Count", "JPX Count", "Total"])
    for strat in STRATEGY_FNS.keys():
        row_data = [strat]
        total    = 0
        for market in ["US", "ASX", "JPX"]:
            c = (all_results.get(market, pd.DataFrame()).get("Best Strategy", pd.Series()) == strat).sum() \
                if market in all_results and not all_results[market].empty else 0
            row_data.append(c); total += c
        row_data.append(total)
        ws_dist.append(row_data)
    _style_header(ws_dist)

    # ── Active Gates tab ──────────────────────────────────────────────────────
    ws_gates = wb.create_sheet(title="🚪 Active Gates")
    ws_gates.append(["ACTIVE ASSET-SPECIFIC GATES"])
    ws_gates.append([f"Loaded from asset_specific_gates.json on run {today}"])
    ws_gates.append([])
    ws_gates.append(["Asset", "Strategy", "Allowed Regimes"])
    if ASSET_SPECIFIC_GATES:
        for asset, strats in ASSET_SPECIFIC_GATES.items():
            for strat, regs in strats.items():
                ws_gates.append([asset, strat, ", ".join(sorted(regs))])
    else:
        ws_gates.append(["—", "—", "(no asset-specific gates; defaults in use)"])
    _style_header(ws_gates, row_num=4)

    wb.save(path)
    return path


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Goofy Phase 5 — Regime-aware screener + position sizing")
    parser.add_argument("--market", choices=["US", "ASX", "JPX", "ALL"],
                        default="ALL")
    args = parser.parse_args()

    today = dt.datetime.now().strftime("%Y-%m-%d")

    loaded_gates = load_asset_gates()
    if loaded_gates:
        print(f"  [Phase 4] {sum(len(v) for v in loaded_gates.values())} "
              f"asset-specific gates loaded across {len(loaded_gates)} assets")
    else:
        print("  [Phase 4] No asset_specific_gates.json — using theory defaults")

    markets_to_run = ["US", "ASX", "JPX"] if args.market == "ALL" else [args.market]

    print(f"""
╔══════════════════════════════════════════════════════════════════════╗
║  GOOFY SCREENER — PHASE 5  |  {today}                    ║
║  Regime-aware verdict  +  Kelly + Vol-Scaled position sizing         ║
║  Markets: {', '.join(markets_to_run):52} ║
║  Train: 2016–2021  |  Test: 2021–{today[:4]}  |  Vol target: {TARGET_VOL*100:.0f}%  ║
╚══════════════════════════════════════════════════════════════════════╝""")

    # ── Step 1: Download ──────────────────────────────────────────────────────
    all_assets = []
    for m in markets_to_run:
        all_assets.extend(UNIVERSE_MAP[m])
    seen = set(); unique_assets = []
    for a in all_assets:
        if a not in seen: unique_assets.append(a); seen.add(a)

    print(f"\n[1/3] Downloading {len(unique_assets)} assets...\n")
    price_data = {}
    ohlc_data  = {}
    for asset in unique_assets:
        try:
            raw = yf.download(asset, start=TRAIN_START, end=TEST_END,
                              auto_adjust=True, progress=False)
            if not raw.empty and len(raw) >= MIN_ROWS:
                if isinstance(raw.columns, pd.MultiIndex):
                    raw.columns = raw.columns.get_level_values(0)
                close = raw["Close"].squeeze()
                if isinstance(close, pd.DataFrame):
                    close = close.iloc[:, 0]
                price_data[asset] = close
                if {"High", "Low", "Close"}.issubset(set(raw.columns)):
                    ohlc_data[asset] = raw[["Open", "High", "Low", "Close"]].copy() \
                        if "Open" in raw.columns else raw[["High", "Low", "Close"]].copy()
                print(f"  ✓ {asset:14} | {len(raw)} rows")
            else:
                print(f"  ✗ {asset:14} | skipped ({len(raw)} rows)")
        except Exception as e:
            print(f"  ✗ {asset:14} | error: {e}")

    print(f"\n  → {len(price_data)} assets ready ({len(ohlc_data)} with OHLC)\n")

    # ── Step 2: Screen ────────────────────────────────────────────────────────
    print("[2/3] Screening + regime + position sizing...\n")
    all_results = {}
    for m in markets_to_run:
        all_results[m] = screen_market(m, UNIVERSE_MAP[m], price_data, ohlc_data)

    # ── Step 3: Save + summarise ──────────────────────────────────────────────
    print("\n[3/3] Saving report...\n")

    all_dfs = [df for df in all_results.values() if not df.empty]
    if not all_dfs:
        print("  No results.")
        return

    combined = pd.concat(all_dfs, ignore_index=True)

    # Console summary
    print(f"\n{'═'*72}")
    print(f"  GOOFY SCREENER — PHASE 5 SUMMARY  |  {today}")
    print(f"{'═'*72}")
    print(f"  Assets screened: {len(combined)}")
    for m in markets_to_run:
        df_m = all_results.get(m, pd.DataFrame())
        if not df_m.empty:
            print(f"    {m:5}: {len(df_m)}")

    print(f"\n  ── Tier Breakdown ──")
    for tier in ["S", "A", "B", "Skip"]:
        icon  = {"S": "⭐", "A": "✅", "B": "🔵", "Skip": "⬜"}.get(tier, "")
        count = len(combined[combined["Tier"] == tier])
        print(f"    {icon} {tier:5}: {count:3}  {'█' * count}")

    print(f"\n  ── Today's Verdict ──")
    for v in ["TRADE", "STAND DOWN", "—"]:
        icon  = {"TRADE": "🟢", "STAND DOWN": "🔴", "—": "⬜"}.get(v, "")
        count = len(combined[combined["Today's Verdict"] == v])
        print(f"    {icon} {v:11}: {count:3}  {'█' * count}")

    print(f"\n  ── Phase 5: Position Sizing (TRADE verdict, S/A/B only) ──")
    trade_list = combined[
        (combined["Today's Verdict"] == "TRADE") &
        (combined["Tier"].isin(["S", "A", "B"]))
    ].sort_values("Score", ascending=False)

    if not trade_list.empty:
        print(f"  {'Asset':14} {'Strategy':18} {'Tier':5} {'Score':6} "
              f"{'Kelly%':8} {'VolScale':9} {'Size%':6} {'Trend':8}")
        print(f"  {'─'*75}")
        for _, r in trade_list.head(20).iterrows():
            k  = f"{r['Kelly %']:.1f}%" if r['Kelly %'] is not None else "  —  "
            vs = f"{r['Vol Scalar']:.2f}"  if r['Vol Scalar'] is not None else "  — "
            sz = f"{r['Recommended Size %']:.0f}%" if r['Recommended Size %'] is not None else " — "
            print(f"  {r['Asset']:14} {r['Best Strategy']:18} {r['Tier']:5} "
                  f"{r['Score']:5.0f} {k:>8} {vs:>9} {sz:>6} {r['Current Trend']}")
    else:
        print("    (no TRADE-verdict S/A/B assets)")

    print(f"\n  ── S-Tier ──")
    s_tier = combined[combined["Tier"] == "S"].sort_values("Score", ascending=False)
    if not s_tier.empty:
        for _, r in s_tier.iterrows():
            sz = (f"{r['Recommended Size %']:.0f}%"
                  if r['Recommended Size %'] is not None else "—")
            v_str = r["Today's Verdict"]
            print(f"    {r['Market']:5} {r['Asset']:14} | {r['Best Strategy']:18} | "
                  f"Sharpe: {r['OUT Sharpe']:5.2f} | Ret: {r['OUT Strat Ret %']:6.1f}% | "
                  f"Score: {r['Score']:.0f}/100 | Verdict: {v_str:11} | "
                  f"Size: {sz}")
    else:
        print("    (none)")

    print(f"\n  ── A-Tier ──")
    a_tier = combined[combined["Tier"] == "A"].sort_values("Score", ascending=False)
    if not a_tier.empty:
        for _, r in a_tier.iterrows():
            sz = (f"{r['Recommended Size %']:.0f}%"
                  if r['Recommended Size %'] is not None else "—")
            v_str = r["Today's Verdict"]
            print(f"    {r['Market']:5} {r['Asset']:14} | {r['Best Strategy']:18} | "
                  f"Sharpe: {r['OUT Sharpe']:5.2f} | Ret: {r['OUT Strat Ret %']:6.1f}% | "
                  f"Score: {r['Score']:.0f}/100 | Verdict: {v_str:11} | "
                  f"Size: {sz}")
    else:
        print("    (none)")

    xlsx_path = write_excel_report(all_results, today)
    print(f"\n  Report saved → {os.path.basename(xlsx_path)}")
    print(f"  Full path:   {xlsx_path}")
    print(f"\n{'═'*72}\n")


if __name__ == "__main__":
    main()
